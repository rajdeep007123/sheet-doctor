#!/usr/bin/env python3
"""
Generates sample-data/messy_sample.xlsx with deliberate data quality problems
for testing excel-doctor.

Run from the repo root:
    python sample-data/generate_xlsx.py

Problems baked in:
  Sheet "Orders"
    - Duplicate header: "customer_id" appears twice (cols B and C)
    - Merged cells: G2:G3 (notes cells merged across two data rows)
    - Formula errors: #DIV/0! (E5), #REF! (E9), #VALUE! (E12)
    - Mixed types in "amount": floats + "N/A" string
    - Empty rows: rows 5 and 10
    - 5 different date formats in "order_date"
    - Single-value column: "status" is always "active"
  Sheet "Summary"
    - Empty sheet (no data)
  Sheet "Archive"
    - Hidden sheet with a small data table
"""

from pathlib import Path
import openpyxl

OUTPUT = Path(__file__).parent / "messy_sample.xlsx"

wb = openpyxl.Workbook()

# ── Sheet 1: Orders ──────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Orders"

# Row 1 — headers (duplicate "customer_id" in cols B and C)
headers = ["order_id", "customer_id", "customer_id", "order_date", "amount", "status", "notes"]
ws.append(headers)

# Rows 2–12 — data with various deliberate problems
# Append all data BEFORE applying merges so row numbering is correct
data = [
    # order_id  cust_id  name       order_date           amount     status    notes
    [1001,      "C001",  "Smith",   "2023-01-15",        250.00,    "active", "First order"],   # row 2
    [1002,      "C002",  "Jones",   "15/02/2023",        180.50,    "active", "Bulk discount"],  # row 3
    [1003,      "C003",  "Brown",   "2023-03-20",        320.00,    "active", ""],               # row 4
    [None,      None,    None,      None,                None,      None,     None],              # row 5 — empty
    [1004,      "C004",  "Wilson",  "March 5, 2023",     "#DIV/0!", "active", "Calc error"],     # row 6
    [1005,      "C005",  "Taylor",  "2023-04-10",        "N/A",     "active", "Pending"],        # row 7
    [1006,      "C006",  "Davis",   "10-05-2023",        415.00,    "active", ""],               # row 8
    [1007,      "C007",  "Miller",  "2023/06/22",        "#REF!",   "active", "Link broken"],    # row 9
    [None,      None,    None,      None,                None,      None,     None],              # row 10 — empty
    [1008,      "C008",  "Moore",   "22 Jul 2023",       "#VALUE!", "active", "Bad data"],       # row 11
    [1009,      "C009",  "Harris",  "2023-09-14",        510.00,    "active", ""],               # row 12
]

for row in data:
    ws.append(row)

# Apply merged cells AFTER data is written
# Merge notes cells for Smith and Jones (G2:G3) — simulates merged data cells
ws.merge_cells("G2:G3")

# ── Sheet 2: Summary (empty) ──────────────────────────────────────────────────
ws_summary = wb.create_sheet("Summary")
# Intentionally left empty

# ── Sheet 3: Archive (hidden) ─────────────────────────────────────────────────
ws_archive = wb.create_sheet("Archive")
ws_archive.append(["order_id", "customer_id", "amount", "archived_date"])
ws_archive.append([999,  "C000", 100.00, "2022-12-31"])
ws_archive.append([998,  "C099",  75.00, "2022-11-15"])
ws_archive.sheet_state = "hidden"

wb.save(OUTPUT)
print(f"Created: {OUTPUT}")
