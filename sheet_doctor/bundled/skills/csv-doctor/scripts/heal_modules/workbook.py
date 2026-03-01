from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from heal_modules.shared import HEADERS, WRITE_ONLY_THRESHOLD, Change, CleanRow, QuarantineRow

def _header_font() -> Font:
    return Font(bold=True, color="FFFFFF")

def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _style_sheet(ws, col_widths: list[int], header_color: str):
    """Apply bold header, color, frozen row, and column widths."""
    fill = _header_fill(header_color)
    font = _header_font()
    for cell in ws[1]:
        cell.font  = font
        cell.fill  = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.freeze_panes = "A2"
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _infer_col_widths(rows: list[list], min_width: int = 10, max_width: int = 60, sample: int = 300) -> list[int]:
    if not rows:
        return []
    widths = [max(min_width, min(max_width, len(str(v)) + 2)) for v in rows[0]]
    for row in rows[1 : sample + 1]:
        for i, val in enumerate(row):
            widths[i] = max(widths[i], min(max_width, len(str(val)) + 2))
    return [max(min_width, min(max_width, w)) for w in widths]

# Accent fills for was_modified / needs_review cells
FILL_MODIFIED = PatternFill("solid", fgColor="FFF2CC")   # soft yellow
FILL_REVIEW   = PatternFill("solid", fgColor="FCE4D6")   # soft orange


def _write_workbook_fast_impl(
    clean_data:  list[CleanRow],
    quarantine:  list[QuarantineRow],
    changelog:   list[Change],
    output_path: Path,
    headers: list[str],
) -> None:
    """Write_only=True path for large outputs — 5–10× faster than regular openpyxl."""
    wb = openpyxl.Workbook(write_only=True)

    hdr_font = Font(bold=True, color="FFFFFF")

    def _hdr_cell(ws, value: str, hex_color: str) -> WriteOnlyCell:
        c = WriteOnlyCell(ws, value=value)
        c.font = hdr_font
        c.fill = PatternFill("solid", fgColor=hex_color)
        return c

    # ── Sheet 1 — Clean Data ────────────────────────────────────────────
    ws1 = wb.create_sheet("Clean Data")
    clean_headers = headers + ["was_modified", "needs_review"]
    for i in range(1, len(clean_headers) + 1):
        ws1.column_dimensions[get_column_letter(i)].width = 15
    ws1.append([_hdr_cell(ws1, h, "4CAF50") for h in clean_headers])
    for entry in clean_data:
        row_out = list(entry.row)
        mod_cell = WriteOnlyCell(ws1, value=entry.was_modified)
        rev_cell = WriteOnlyCell(ws1, value=entry.needs_review)
        if entry.was_modified:
            mod_cell.fill = FILL_MODIFIED
        if entry.needs_review:
            rev_cell.fill = FILL_REVIEW
        ws1.append(row_out + [mod_cell, rev_cell])

    # ── Sheet 2 — Quarantine ────────────────────────────────────────────
    ws2 = wb.create_sheet("Quarantine")
    quarantine_headers = headers + ["quarantine_reason"]
    for i in range(1, len(quarantine_headers) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 15
    ws2.append([_hdr_cell(ws2, h, "E53935") for h in quarantine_headers])
    for q in quarantine:
        ws2.append(q.row + [q.reason])

    # ── Sheet 3 — Change Log ────────────────────────────────────────────
    ws3 = wb.create_sheet("Change Log")
    log_headers = ["original_row_number", "column_affected",
                   "original_value", "new_value", "action_taken", "reason"]
    for i in range(1, len(log_headers) + 1):
        ws3.column_dimensions[get_column_letter(i)].width = 15
    ws3.append([_hdr_cell(ws3, h, "1565C0") for h in log_headers])
    for c in changelog:
        ws3.append([c.original_row_number, c.column_affected,
                    c.original_value, c.new_value, c.action_taken, c.reason])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _write_workbook_standard_impl(
    clean_data:  list[CleanRow],
    quarantine:  list[QuarantineRow],
    changelog:   list[Change],
    output_path: Path,
    headers: list[str] | None = None,
) -> None:
    wb = openpyxl.Workbook()

    # ── Sheet 1 — Clean Data ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Clean Data"
    clean_headers = headers + ["was_modified", "needs_review"]
    clean_rows_for_width = [clean_headers]
    ws1.append(clean_headers)
    for entry in clean_data:
        row_out = entry.row + [entry.was_modified, entry.needs_review]
        ws1.append(row_out)
        clean_rows_for_width.append(row_out)
        # Accent modified / review flag cells
        last = ws1.max_row
        mod_cell    = ws1.cell(last, len(headers) + 1)
        review_cell = ws1.cell(last, len(headers) + 2)
        if entry.was_modified:
            mod_cell.fill = FILL_MODIFIED
        if entry.needs_review:
            review_cell.fill = FILL_REVIEW
    _style_sheet(ws1, _infer_col_widths(clean_rows_for_width), "4CAF50")   # green

    # ── Sheet 2 — Quarantine ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Quarantine")
    quarantine_headers = headers + ["quarantine_reason"]
    quarantine_rows_for_width = [quarantine_headers]
    ws2.append(quarantine_headers)
    for q in quarantine:
        row_out = q.row + [q.reason]
        ws2.append(row_out)
        quarantine_rows_for_width.append(row_out)
    _style_sheet(ws2, _infer_col_widths(quarantine_rows_for_width), "E53935")   # red

    # ── Sheet 3 — Change Log ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Change Log")
    log_headers = ["original_row_number", "column_affected",
                   "original_value", "new_value", "action_taken", "reason"]
    log_rows_for_width = [log_headers]
    ws3.append(log_headers)
    for c in changelog:
        row_out = [c.original_row_number, c.column_affected,
                   c.original_value, c.new_value, c.action_taken, c.reason]
        ws3.append(row_out)
        log_rows_for_width.append(row_out)
    _style_sheet(ws3, _infer_col_widths(log_rows_for_width), "1565C0")   # blue

    # Notes column text-wrap in Sheet 1
    if "Notes" in headers:
        notes_col = get_column_letter(headers.index("Notes") + 1)
        for cell in ws1[notes_col][1:]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Reason column text-wrap in Sheet 3
    for cell in ws3["F"][1:]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
