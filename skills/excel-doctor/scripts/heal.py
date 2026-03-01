#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import tempfile
import zipfile
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parents[2]
sys.path.insert(0, str(ROOT_DIR))

from sheet_doctor import __version__ as TOOL_VERSION
from sheet_doctor.contracts import build_contract, build_run_summary

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("ERROR: openpyxl not installed — run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SMART_QUOTES = {
    "\u201c": '"',
    "\u201d": '"',
    "\u2018": "'",
    "\u2019": "'",
}
INTERNAL_SHEETS = {"Change Log"}
HEADER_HINT_RE = re.compile(r"[A-Za-z]")


@dataclass
class Change:
    sheet: str
    cell_or_range: str
    action: str
    old_value: str
    new_value: str
    reason: str


def is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def is_sheet_empty(sheet) -> bool:
    for row in sheet.iter_rows(values_only=True):
        if any(not is_blank(value) for value in row):
            return False
    return True


def text(value) -> str:
    if value is None:
        return ""
    return str(value)


def is_encrypted_ooxml(file_path: Path) -> bool:
    if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    try:
        with zipfile.ZipFile(file_path) as archive:
            names = set(archive.namelist())
    except zipfile.BadZipFile:
        return False
    return {"EncryptedPackage", "EncryptionInfo"}.issubset(names)


def row_values(sheet, row_idx: int) -> list:
    return [sheet.cell(row=row_idx, column=col).value for col in range(1, sheet.max_column + 1)]


def non_empty_texts(values: list) -> list[str]:
    return [text(v).strip() for v in values if not is_blank(v)]


def headerish_score(values: list) -> int:
    score = 0
    for value in values:
        if is_blank(value):
            continue
        raw = text(value).strip()
        if not raw:
            continue
        if isinstance(value, (int, float, datetime)):
            continue
        if raw.startswith("="):
            continue
        if HEADER_HINT_RE.search(raw):
            score += 1
    return score


def detect_header_band(sheet) -> dict:
    search_end = min(sheet.max_row, 8)
    candidates = []
    for row_idx in range(1, search_end + 1):
        values = row_values(sheet, row_idx)
        if len(non_empty_texts(values)) < 2:
            continue
        score = headerish_score(values)
        if score >= 2:
            candidates.append((row_idx, score))
    if not candidates:
        return {"header_row": 1, "header_band_rows": [1], "metadata_rows": []}
    header_row = max(candidates, key=lambda item: (item[1], item[0]))[0]
    band_start = header_row
    while band_start > 1:
        values = row_values(sheet, band_start - 1)
        if len(non_empty_texts(values)) < 2 or headerish_score(values) < 2:
            break
        band_start -= 1
    return {
        "header_row": header_row,
        "header_band_rows": list(range(band_start, header_row + 1)),
        "metadata_rows": list(range(1, band_start)),
    }


def merge_header_band_rows(sheet, rows: list[int]) -> list[str]:
    width = sheet.max_column
    merged = []
    for col in range(1, width + 1):
        tokens = []
        seen = set()
        current = ""
        for row in rows:
            value = text(sheet.cell(row=row, column=col).value).strip()
            if value:
                current = value
            elif current:
                value = current
            if not value:
                continue
            key = value.lower()
            if key in seen:
                continue
            seen.add(key)
            tokens.append(value)
        merged.append(" ".join(tokens).strip())
    return merged


def clean_text(value: str) -> tuple[str, list[str]]:
    new_value = value
    reasons = []
    if "\ufeff" in new_value:
        new_value = new_value.replace("\ufeff", "")
        reasons.append("BOM removed")
    if "\x00" in new_value:
        new_value = new_value.replace("\x00", "")
        reasons.append("NULL byte removed")
    if "\r" in new_value or "\n" in new_value:
        new_value = new_value.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
        reasons.append("line breaks replaced with spaces")
    had_smart = any(char in new_value for char in SMART_QUOTES)
    for smart, straight in SMART_QUOTES.items():
        new_value = new_value.replace(smart, straight)
    if had_smart:
        reasons.append("smart quotes normalised")
    collapsed = " ".join(new_value.split())
    if collapsed != new_value:
        reasons.append("extra whitespace collapsed")
    return collapsed, reasons


def normalise_date_text(value: str) -> tuple[str, bool, str]:
    v = value.strip()
    if not v or re.match(r"^\d{4}-\d{2}-\d{2}$", v):
        return v, False, ""
    for fmt, reason in [
        ("%Y/%m/%d", "YYYY/MM/DD normalised to YYYY-MM-DD"),
        ("%B %d %Y", "Month D YYYY normalised to YYYY-MM-DD"),
        ("%b %d %Y", "Mon D YYYY normalised to YYYY-MM-DD"),
        ("%B %d, %Y", "Month D, YYYY normalised to YYYY-MM-DD"),
        ("%b %d, %Y", "Mon D, YYYY normalised to YYYY-MM-DD"),
    ]:
        try:
            dt = datetime.strptime(v, fmt)
            return dt.strftime("%Y-%m-%d"), True, reason
        except ValueError:
            pass
    for pattern, preferred in [
        (r"^(\d{1,2})/(\d{1,2})/(\d{4})$", "slash"),
        (r"^(\d{1,2})-(\d{1,2})-(\d{4})$", "dash"),
        (r"^(\d{2})-(\d{2})-(\d{2})$", "short_dash"),
    ]:
        m = re.match(pattern, v)
        if not m:
            continue
        a, b, c = (int(m.group(i)) for i in range(1, 4))
        if preferred == "short_dash":
            year = 2000 + c if c < 50 else 1900 + c
            pairs = [(a, b, "DD-MM-YY"), (b, a, "MM-DD-YY")]
        else:
            year = c
            pairs = [(a, b, "DD/MM/YYYY" if preferred == "slash" else "DD-MM-YYYY"), (b, a, "MM/DD/YYYY" if preferred == "slash" else "MM-DD-YYYY")]
        for day, month, label in pairs:
            try:
                dt = datetime(year, month, day)
                return dt.strftime("%Y-%m-%d"), True, f"{label} normalised to YYYY-MM-DD (day-first preferred when ambiguous)"
            except ValueError:
                continue
    return v, False, ""


def normalise_header(raw_header, index: int) -> str:
    base = text(raw_header).strip()
    base = " ".join(base.split())
    if not base:
        return f"column_{index}"
    return base


def add_change(changes: list[Change], sheet_name: str, target: str, action: str, old_value, new_value, reason: str) -> None:
    changes.append(Change(sheet=sheet_name, cell_or_range=target, action=action, old_value=text(old_value), new_value=text(new_value), reason=reason))


def trim_empty_edge_columns(sheet, changes: list[Change], stats: Counter) -> None:
    if sheet.max_column == 0:
        return

    def data_non_empty(col_idx: int) -> int:
        return sum(1 for row_idx in range(1, sheet.max_row + 1) if not is_blank(sheet.cell(row=row_idx, column=col_idx).value))

    leading = 0
    for col_idx in range(1, sheet.max_column + 1):
        if data_non_empty(col_idx) > 0:
            break
        leading += 1
    trailing = 0
    for col_idx in range(sheet.max_column, 0, -1):
        if data_non_empty(col_idx) > 0:
            break
        trailing += 1

    if trailing:
        start = sheet.max_column - trailing + 1
        sheet.delete_cols(start, trailing)
        stats["edge_columns_trimmed"] += trailing
        add_change(changes, sheet.title, f"{start}:{start + trailing - 1}", "Removed", "[empty edge columns]", "", "Trailing empty workbook columns removed")
    if leading:
        sheet.delete_cols(1, leading)
        stats["edge_columns_trimmed"] += leading
        add_change(changes, sheet.title, f"1:{leading}", "Removed", "[empty edge columns]", "", "Leading empty workbook columns removed")


def heal_sheet(sheet, changes: list[Change], stats: Counter) -> None:
    sheet_name = sheet.title
    if is_sheet_empty(sheet):
        stats["empty_sheets_skipped"] += 1
        return

    plan = detect_header_band(sheet)
    if plan["metadata_rows"]:
        first = plan["metadata_rows"][0]
        count = len(plan["metadata_rows"])
        sheet.delete_rows(first, count)
        stats["metadata_rows_removed"] += count
        add_change(changes, sheet_name, f"{first}:{first + count - 1}", "Removed", "[metadata rows]", "", "Workbook preamble rows removed before table header")
        plan = detect_header_band(sheet)

    if len(plan["header_band_rows"]) > 1:
        merged_header = merge_header_band_rows(sheet, plan["header_band_rows"])
        for col_idx, value in enumerate(merged_header, start=1):
            sheet.cell(row=1, column=col_idx).value = value
        delete_start = 2
        delete_count = len(plan["header_band_rows"]) - 1
        sheet.delete_rows(delete_start, delete_count)
        stats["header_bands_flattened"] += 1
        add_change(changes, sheet_name, f"1:{len(plan['header_band_rows'])}", "Fixed", "[stacked header rows]", "[flattened header row]", "Stacked workbook header band flattened into one header row")

    merged_ranges = list(sheet.merged_cells.ranges)
    for merged in merged_ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        range_ref = str(merged)
        anchor_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(range_ref)
        stats["merged_ranges_unmerged"] += 1
        add_change(changes, sheet_name, range_ref, "Fixed", "[merged]", "[unmerged]", "Merged range unmerged for tabular compatibility")
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row == min_row and col == min_col:
                    continue
                cell = sheet.cell(row=row, column=col)
                old = cell.value
                if old != anchor_value:
                    cell.value = anchor_value
                    stats["cells_filled_from_merged_anchor"] += 1
                    add_change(changes, sheet_name, cell.coordinate, "Fixed", old, anchor_value, f"Filled from merged anchor cell {sheet.cell(min_row, min_col).coordinate}")

    trim_empty_edge_columns(sheet, changes, stats)

    max_col = sheet.max_column
    if max_col == 0:
        return

    seen_headers = Counter()
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=1, column=col)
        old = cell.value
        header = normalise_header(old, col)
        key = header.lower()
        seen_headers[key] += 1
        if seen_headers[key] > 1:
            deduped = f"{header}_{seen_headers[key]}"
            stats["headers_deduplicated"] += 1
            reason = "Duplicate header renamed with numeric suffix"
        else:
            deduped = header
            reason = "Header standardised"
        if text(old) != deduped:
            cell.value = deduped
            stats["headers_standardised"] += 1
            add_change(changes, sheet_name, cell.coordinate, "Fixed", old, deduped, reason)

    for row in range(2, sheet.max_row + 1):
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            value = cell.value
            if value is None:
                continue
            if cell.data_type == "f":
                stats["formula_cells_preserved"] += 1
                continue
            if not isinstance(value, str):
                continue
            cleaned, reasons = clean_text(value)
            date_value, date_changed, date_reason = normalise_date_text(cleaned)
            final = date_value if date_changed else cleaned
            if date_changed:
                reasons.append(date_reason)
            if final != value:
                cell.value = final
                if reasons:
                    stats["text_cells_cleaned"] += 1
                    if date_changed:
                        stats["dates_normalised"] += 1
                    add_change(changes, sheet_name, cell.coordinate, "Fixed", value, final, "; ".join(reasons))

    for row in range(sheet.max_row, 1, -1):
        if all(is_blank(v) for v in row_values(sheet, row)):
            sheet.delete_rows(row, 1)
            stats["empty_rows_removed"] += 1
            add_change(changes, sheet_name, f"{row}:{row}", "Removed", "[empty row]", "", "Fully empty row removed")


def write_change_log_sheet(workbook, changes: list[Change]) -> None:
    if "Change Log" in workbook.sheetnames:
        workbook.remove(workbook["Change Log"])
    ws = workbook.create_sheet("Change Log")
    ws.append(["sheet", "cell_or_range", "action", "old_value", "new_value", "reason"])
    for change in changes:
        ws.append([change.sheet, change.cell_or_range, change.action, change.old_value, change.new_value, change.reason])


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Heal .xlsx/.xlsm workbooks and append a Change Log sheet.")
    parser.add_argument("input", help="Input .xlsx/.xlsm file")
    parser.add_argument("output", nargs="?", help="Output workbook path")
    parser.add_argument("--json-summary", dest="json_summary", help="Optional path to write a structured JSON healing summary for UI/backend use")
    return parser.parse_args(argv)


def execute_healing(input_path: Path, output_path: Path) -> tuple[list[Change], Counter]:
    if is_encrypted_ooxml(input_path):
        raise ValueError("Password-protected / encrypted OOXML workbooks are not supported")
    keep_vba = input_path.suffix.lower() == ".xlsm"
    try:
        workbook = load_workbook(input_path, keep_vba=keep_vba)
    except Exception as exc:
        raise ValueError(f"Could not read workbook: {exc}") from exc

    changes: list[Change] = []
    stats = Counter()
    for sheet in workbook.worksheets:
        if sheet.title in INTERNAL_SHEETS:
            continue
        stats["sheets_processed"] += 1
        heal_sheet(sheet, changes, stats)

    write_change_log_sheet(workbook, changes)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=f".{output_path.stem}.", suffix=output_path.suffix, dir=str(output_path.parent))
    os.close(fd)
    temp_path = Path(tmp_name)
    try:
        workbook.save(temp_path)
        os.replace(temp_path, output_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()
    return changes, stats


def build_structured_summary(*, input_path: Path, output_path: Path, changes: list[Change], stats: Counter) -> dict:
    contract = build_contract("excel_doctor.heal_summary")
    warnings = []
    if stats["formula_cells_preserved"] > 0:
        warnings.append("Formula cells were preserved unchanged. excel-doctor does not recalculate formulas.")
    if input_path.suffix.lower() == ".xlsm":
        warnings.append(".xlsm macro preservation only holds while the output remains .xlsm.")
    return {
        "contract": contract,
        "schema_version": contract["version"],
        "tool_version": TOOL_VERSION,
        "input_file": str(input_path),
        "output_file": str(output_path),
        "mode": "workbook-native",
        "stats": dict(stats),
        "changes_logged": len(changes),
        "warnings": warnings,
        "assumptions": [
            "Workbook-native healing preserves workbook sheets instead of flattening them into a CSV-style table",
            "Ambiguous DD/MM vs MM/DD dates prefer day-first when both parse",
            "Formula cells are preserved unchanged; formula cache values are not reconstructed",
            ".xlsm macros are preserved only when the workbook stays in .xlsm format",
        ],
        "run_summary": build_run_summary(
            tool="excel-doctor",
            script="heal.py",
            input_path=input_path,
            output_path=output_path,
            metrics={
                "sheets_processed": stats["sheets_processed"],
                "changes_logged": len(changes),
                "headers_standardised": stats["headers_standardised"],
                "headers_deduplicated": stats["headers_deduplicated"],
                "header_bands_flattened": stats["header_bands_flattened"],
                "metadata_rows_removed": stats["metadata_rows_removed"],
                "merged_ranges_unmerged": stats["merged_ranges_unmerged"],
                "formula_cells_preserved": stats["formula_cells_preserved"],
                "dates_normalised": stats["dates_normalised"],
                "empty_rows_removed": stats["empty_rows_removed"],
                "edge_columns_trimmed": stats["edge_columns_trimmed"],
                "mode": "workbook-native",
            },
        ),
    }


def main():
    args = parse_args(sys.argv[1:])
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: File not found: {input_path}", file=sys.stderr)
        sys.exit(1)
    suffix = input_path.suffix.lower()
    if suffix == ".xls":
        print("ERROR: .xls is not supported by excel-doctor. Use csv-doctor tabular rescue or convert to .xlsx first.", file=sys.stderr)
        sys.exit(1)
    if suffix not in (".xlsx", ".xlsm"):
        print(f"ERROR: Expected an .xlsx/.xlsm file, got: {input_path.suffix}", file=sys.stderr)
        sys.exit(1)
    output_path = Path(args.output) if args.output else input_path.with_name(f"{input_path.stem}_healed{input_path.suffix}")
    try:
        changes, stats = execute_healing(input_path, output_path)
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
    if args.json_summary:
        summary_path = Path(args.json_summary)
        summary_path.parent.mkdir(parents=True, exist_ok=True)
        summary_path.write_text(json.dumps(build_structured_summary(input_path=input_path, output_path=output_path, changes=changes, stats=stats), indent=2, ensure_ascii=False), encoding="utf-8")
    print()
    print("═" * 62)
    print("  Excel Doctor  ·  Heal Report")
    print("═" * 62)
    print(f"  Input file   : {input_path.name}")
    print(f"  Output file  : {output_path.name}")
    print("  Mode         : workbook-native")
    print("─" * 62)
    print(f"  Sheets processed                : {stats['sheets_processed']}")
    print(f"  Changes logged                  : {len(changes)}")
    print(f"    · Headers standardised        : {stats['headers_standardised']}")
    print(f"    · Headers deduplicated        : {stats['headers_deduplicated']}")
    print(f"    · Header bands flattened      : {stats['header_bands_flattened']}")
    print(f"    · Metadata rows removed       : {stats['metadata_rows_removed']}")
    print(f"    · Merged ranges unmerged      : {stats['merged_ranges_unmerged']}")
    print(f"    · Cells filled from merges    : {stats['cells_filled_from_merged_anchor']}")
    print(f"    · Formula cells preserved     : {stats['formula_cells_preserved']}")
    print(f"    · Text cells cleaned          : {stats['text_cells_cleaned']}")
    print(f"    · Dates normalised            : {stats['dates_normalised']}")
    print(f"    · Empty rows removed          : {stats['empty_rows_removed']}")
    print(f"    · Edge columns trimmed        : {stats['edge_columns_trimmed']}")
    print("─" * 62)
    if stats["formula_cells_preserved"] > 0:
        print("  WARNINGS:")
        print("    · Formula cells were preserved unchanged; excel-doctor does not recalculate formulas")
        if input_path.suffix.lower() == ".xlsm":
            print("    · .xlsm macro preservation only holds while the output remains .xlsm")
        print("─" * 62)
    print("  ASSUMPTIONS:")
    print("    · Workbook-native healing preserves workbook sheets instead of flattening them into a CSV-style table")
    print("    · Ambiguous DD/MM vs MM/DD dates prefer day-first when both parse")
    print("    · Formula cells are preserved unchanged; formula cache values are not reconstructed")
    print("    · .xlsm macros are preserved only when the workbook stays in .xlsm format")
    print("═" * 62)
    print()


if __name__ == "__main__":
    main()
