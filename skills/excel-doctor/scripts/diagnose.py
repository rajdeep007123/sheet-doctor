#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import sys
import zipfile
from collections import Counter
from datetime import date, datetime, time
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parents[2]
sys.path.insert(0, str(ROOT_DIR))

from sheet_doctor import __version__ as TOOL_VERSION
from sheet_doctor.contracts import build_contract, build_run_summary

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl not installed — run: pip install openpyxl"}), file=sys.stdout)
    sys.exit(1)

ERROR_VALUES = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#N/A", "#NUM!"}
INTERNAL_SHEETS = {"Change Log"}
STRUCTURAL_ROW_RE = re.compile(r"^\s*(grand\s+total|subtotal|sub-total|total|sum)\b", re.IGNORECASE)
DATE_LIKE_RE = re.compile(r".*\d{2,4}[-/]\d{1,2}[-/]\d{1,4}.*|.*\d{1,2}\s+\w{3,9}\s+\d{2,4}.*")
DATE_PATTERNS = [
    (r"^\d{4}-\d{2}-\d{2}$", "YYYY-MM-DD"),
    (r"^\d{2}/\d{2}/\d{4}$", "DD/MM/YYYY or MM/DD/YYYY"),
    (r"^\d{2}-\d{2}-\d{4}$", "DD-MM-YYYY or MM-DD-YYYY"),
    (r"^\d{2}/\d{2}/\d{2}$", "DD/MM/YY or MM/DD/YY"),
    (r"^\d{2}-\d{2}-\d{2}$", "DD-MM-YY or MM-DD-YY"),
    (r"^\d{1,2}\s+\w+\s+\d{4}$", "D Month YYYY"),
    (r"^\w+\s+\d{1,2},?\s+\d{4}$", "Month D, YYYY"),
    (r"^\d{8}$", "YYYYMMDD"),
    (r"^\d{4}/\d{2}/\d{2}$", "YYYY/MM/DD"),
    (r"^\d{1,2}/\d{1,2}/\d{4}$", "M/D/YYYY or D/M/YYYY"),
]
NOTES_ROW_RE = re.compile(r"\b(note|comment|memo|generated|approved|review|report)\b", re.IGNORECASE)
HEADER_HINT_RE = re.compile(r"[A-Za-z]")


def is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value).strip()


def is_encrypted_ooxml(file_path: Path) -> bool:
    if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    try:
        with zipfile.ZipFile(file_path) as archive:
            names = set(archive.namelist())
    except zipfile.BadZipFile:
        return False
    return {"EncryptedPackage", "EncryptionInfo"}.issubset(names)


def classify_value(value) -> str | None:
    if is_blank(value):
        return None
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, (datetime, date, time)):
        return "date"
    text = to_text(value)
    if text.upper() in ERROR_VALUES:
        return "error"
    return "text"


def is_sheet_empty(sheet) -> bool:
    for row in sheet.iter_rows(values_only=True):
        if any(not is_blank(value) for value in row):
            return False
    return True


def row_values(sheet, row_idx: int) -> list:
    return [sheet.cell(row=row_idx, column=col).value for col in range(1, sheet.max_column + 1)]


def non_empty_texts(values: list) -> list[str]:
    return [to_text(value) for value in values if not is_blank(value)]


def headerish_score(values: list) -> int:
    score = 0
    for value in values:
        if is_blank(value):
            continue
        text = to_text(value)
        if classify_value(value) in {"number", "date", "error"}:
            continue
        if HEADER_HINT_RE.search(text):
            score += 1
    return score


def detect_header_band(sheet) -> dict:
    search_end = min(sheet.max_row, 8)
    candidates = []
    for row_idx in range(1, search_end + 1):
        values = row_values(sheet, row_idx)
        non_empty = non_empty_texts(values)
        if len(non_empty) < 2:
            continue
        score = headerish_score(values)
        next_score = headerish_score(row_values(sheet, row_idx + 1)) if row_idx < sheet.max_row else 0
        if score >= 2:
            candidates.append((row_idx, score, next_score))
    if not candidates:
        return {"header_row": 1, "header_band_rows": [1], "metadata_rows": []}

    header_row = max(candidates, key=lambda item: (item[1], item[2], item[0]))[0]
    band_start = header_row
    while band_start > 1:
        values = row_values(sheet, band_start - 1)
        non_empty = non_empty_texts(values)
        if len(non_empty) < 2:
            break
        if headerish_score(values) < 2:
            break
        band_start -= 1
    return {
        "header_row": header_row,
        "header_band_rows": list(range(band_start, header_row + 1)),
        "metadata_rows": list(range(1, band_start)),
    }


def headers_for_sheet(sheet, header_row: int = 1) -> list[str]:
    headers = []
    for col_idx in range(1, sheet.max_column + 1):
        raw = sheet.cell(row=header_row, column=col_idx).value
        cleaned = to_text(raw)
        headers.append(cleaned if cleaned else f"column_{col_idx}")
    return headers


def header_whitespace(headers: list[str]) -> list[str]:
    return [header for header in headers if header != header.strip()]


def duplicate_headers(headers: list[str]) -> list[str]:
    canonical_to_name = {}
    counts = Counter()
    for header in headers:
        key = header.strip().lower()
        if not key:
            continue
        counts[key] += 1
        canonical_to_name.setdefault(key, header.strip())
    return [canonical_to_name[key] for key, n in counts.items() if n > 1]


def detect_date_format(value: str) -> str | None:
    for pattern, label in DATE_PATTERNS:
        if re.match(pattern, value):
            return label
    return None


def detect_mixed_date_formats(text_values: list[str]) -> dict:
    format_examples: dict[str, str] = {}
    looks_like_date = 0
    for value in text_values:
        if DATE_LIKE_RE.match(value):
            looks_like_date += 1
        label = detect_date_format(value)
        if label and label not in format_examples:
            format_examples[label] = value
    if looks_like_date < 2 or len(format_examples) <= 1:
        return {}
    return {"formats_found": list(format_examples.keys()), "examples": format_examples}


def scan_formula_errors(values_sheet) -> list[dict]:
    errors = []
    for row in values_sheet.iter_rows():
        for cell in row:
            if is_blank(cell.value):
                continue
            if cell.data_type == "e":
                errors.append({"cell": cell.coordinate, "value": to_text(cell.value)})
                continue
            if isinstance(cell.value, str) and cell.value.strip().upper() in ERROR_VALUES:
                errors.append({"cell": cell.coordinate, "value": cell.value.strip()})
    return errors


def scan_formula_cells(formula_sheet) -> list[dict]:
    formulas = []
    for row in formula_sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append({"cell": cell.coordinate, "formula": cell.value})
    return formulas


def scan_formula_cache_misses(formula_sheet, values_sheet) -> list[dict]:
    misses = []
    for row in formula_sheet.iter_rows():
        for formula_cell in row:
            value = formula_cell.value
            if not isinstance(value, str) or not value.startswith("="):
                continue
            cached_cell = values_sheet[formula_cell.coordinate]
            if cached_cell.value is None:
                misses.append({
                    "cell": formula_cell.coordinate,
                    "formula": value,
                    "reason": "No cached value. Open/recalculate in Excel and save.",
                })
    return misses


def scan_empty_rows(sheet, data_start_row: int) -> list[int]:
    empty = []
    for row_idx in range(max(2, data_start_row), sheet.max_row + 1):
        if all(is_blank(value) for value in row_values(sheet, row_idx)):
            empty.append(row_idx)
    return empty


def scan_notes_rows(sheet, data_start_row: int) -> list[dict]:
    notes = []
    for row_idx in range(max(2, data_start_row), sheet.max_row + 1):
        values = row_values(sheet, row_idx)
        non_empty = non_empty_texts(values)
        if len(non_empty) != 1:
            continue
        text = non_empty[0]
        if len(text) > 50 and len(text.split()) >= 8 and NOTES_ROW_RE.search(text):
            notes.append({"row": row_idx, "text": text[:120]})
    return notes


def scan_structural_rows(sheet, data_start_row: int) -> list[dict]:
    structural = []
    for row_idx in range(max(2, data_start_row), sheet.max_row + 1):
        values = row_values(sheet, row_idx)
        non_empty = non_empty_texts(values)
        if not non_empty:
            continue
        first = non_empty[0]
        if STRUCTURAL_ROW_RE.match(first) and len(non_empty) <= 3:
            structural.append({"row": row_idx, "label": first})
    return structural


def scan_empty_edge_columns(sheet, header_row: int) -> dict:
    leading = []
    trailing = []
    cols = list(range(1, sheet.max_column + 1))
    if not cols:
        return {"leading": [], "trailing": []}

    def data_non_empty(col_idx: int) -> int:
        return sum(1 for row_idx in range(header_row + 1, sheet.max_row + 1) if not is_blank(sheet.cell(row=row_idx, column=col_idx).value))

    for col_idx in cols:
        header_text = to_text(sheet.cell(row=header_row, column=col_idx).value)
        if header_text or data_non_empty(col_idx) > 0:
            break
        leading.append(col_idx)

    for col_idx in reversed(cols):
        header_text = to_text(sheet.cell(row=header_row, column=col_idx).value)
        if header_text or data_non_empty(col_idx) > 0:
            break
        trailing.append(col_idx)

    return {"leading": leading, "trailing": list(reversed(trailing))}


def scan_columns(sheet, headers: list[str], data_start_row: int) -> tuple[dict, list[str], dict, dict, dict]:
    mixed_types = {}
    empty_columns = []
    single_value_columns = {}
    date_formats = {}
    high_null_columns = {}
    data_row_count = max(sheet.max_row - data_start_row + 1, 0)

    for col_idx, header in enumerate(headers, start=1):
        types_seen = set()
        type_examples = {}
        non_empty_text = []
        canonical_values = []
        empty_cells = 0
        for row_idx in range(data_start_row, sheet.max_row + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            value_type = classify_value(value)
            if not value_type:
                empty_cells += 1
                continue
            types_seen.add(value_type)
            type_examples.setdefault(value_type, to_text(value))
            text = to_text(value)
            canonical_values.append(text)
            if value_type == "text":
                non_empty_text.append(text)
        if not canonical_values:
            empty_columns.append(header)
            continue
        unique_values = {value for value in canonical_values if value != ""}
        if len(unique_values) == 1:
            single_value_columns[header] = next(iter(unique_values))
        if len(types_seen) > 1:
            mixed_types[header] = {"types": sorted(types_seen), "examples": {t: type_examples[t] for t in sorted(type_examples)}}
        date_mix = detect_mixed_date_formats(non_empty_text)
        if date_mix:
            date_formats[header] = date_mix
        if data_row_count >= 5:
            null_ratio = round(empty_cells / data_row_count, 2)
            if 0.8 <= null_ratio < 1.0:
                high_null_columns[header] = {"null_ratio": null_ratio, "empty_cells": empty_cells, "data_rows": data_row_count}
    return mixed_types, empty_columns, single_value_columns, date_formats, high_null_columns


def count_issue_events(report: dict) -> dict:
    sheets = report["sheets"]
    return {
        "empty_sheets": len(sheets.get("empty", [])),
        "hidden_sheets": len(sheets.get("hidden", [])),
        "very_hidden_sheets": len(sheets.get("very_hidden", [])),
        "merged_ranges": sum(len(v) for v in report.get("merged_cells", {}).values()),
        "formula_cells": sum(len(v) for v in report.get("formula_cells", {}).values()),
        "formula_errors": sum(len(v) for v in report.get("formula_errors", {}).values()),
        "formula_cache_misses": sum(len(v) for v in report.get("formula_cache_misses", {}).values()),
        "mixed_type_columns": sum(len(v) for v in report.get("mixed_types", {}).values()),
        "empty_rows": sum(v["count"] for v in report.get("empty_rows", {}).values()),
        "empty_columns": sum(len(v) for v in report.get("empty_columns", {}).values()),
        "duplicate_headers": sum(len(v) for v in report.get("duplicate_headers", {}).values()),
        "whitespace_headers": sum(len(v) for v in report.get("whitespace_headers", {}).values()),
        "date_format_columns": sum(len(v) for v in report.get("date_formats", {}).values()),
        "single_value_columns": sum(len(v) for v in report.get("single_value_columns", {}).values()),
        "structural_rows": sum(len(v) for v in report.get("structural_rows", {}).values()),
        "notes_rows": sum(len(v) for v in report.get("notes_rows", {}).values()),
        "high_null_columns": sum(len(v) for v in report.get("high_null_columns", {}).values()),
        "header_bands": sum(1 for v in report.get("header_bands", {}).values() if len(v.get("rows", [])) > 1),
        "metadata_rows": sum(len(v) for v in report.get("metadata_rows", {}).values()),
        "empty_edge_columns": sum(len(v.get("leading", [])) + len(v.get("trailing", [])) for v in report.get("empty_edge_columns", {}).values()),
    }


def build_summary(report: dict) -> dict:
    counts = count_issue_events(report)
    critical = 0
    high = 0
    medium = 0
    if counts["formula_errors"] > 0:
        critical += 1
    if counts["formula_cache_misses"] > 0:
        high += 1
    if counts["duplicate_headers"] > 0:
        high += 1
    if counts["mixed_type_columns"] > 0:
        high += 1
    if counts["date_format_columns"] > 0:
        high += 1
    if counts["merged_ranges"] > 0:
        high += 1
    if counts["header_bands"] > 0 or counts["metadata_rows"] > 0:
        high += 1
    if counts["empty_sheets"] > 0:
        medium += 1
    if counts["hidden_sheets"] > 0 or counts["very_hidden_sheets"] > 0:
        medium += 1
    if counts["empty_rows"] > 0:
        medium += 1
    if counts["empty_columns"] > 0 or counts["empty_edge_columns"] > 0:
        medium += 1
    if counts["single_value_columns"] > 0:
        medium += 1
    if counts["whitespace_headers"] > 0:
        medium += 1
    if counts["structural_rows"] > 0 or counts["notes_rows"] > 0:
        medium += 1
    if counts["high_null_columns"] > 0:
        medium += 1
    if critical > 0 or high >= 3:
        verdict = "CRITICAL"
    elif high > 0 or medium >= 2:
        verdict = "NEEDS ATTENTION"
    else:
        verdict = "HEALTHY"
    issue_categories_triggered = sum(1 for _, v in counts.items() if v > 0)
    issue_count = sum(counts.values())
    return {
        "verdict": verdict,
        "issue_count": issue_count,
        "issue_categories_triggered": issue_categories_triggered,
        "severity_breakdown": {
            "critical_categories": critical,
            "high_categories": high,
            "medium_categories": medium,
        },
    }


def build_manual_review_warnings(report: dict) -> list[str]:
    counts = report["issue_counts"]
    warnings = []
    if counts["formula_cells"] > 0:
        warnings.append("Formula cells are present. excel-doctor preserves formulas and does not recalculate business logic.")
    if counts["formula_errors"] > 0:
        warnings.append("Formula error cells still need manual spreadsheet review. excel-doctor does not repair broken formulas.")
    if counts["formula_cache_misses"] > 0:
        warnings.append("Some formula cells have no cached values. Reopen the workbook in Excel and save after recalculation if the cached results matter.")
    if counts["hidden_sheets"] > 0 or counts["very_hidden_sheets"] > 0:
        warnings.append("Hidden sheets exist. Workbook context may depend on tabs the user cannot currently see.")
    if counts["header_bands"] > 0 or counts["metadata_rows"] > 0:
        warnings.append("Header-band and metadata-row detection is heuristic. Confirm the detected table starts before trusting downstream analysis.")
    return warnings


def build_report(file_path: Path) -> dict:
    if is_encrypted_ooxml(file_path):
        raise ValueError("Password-protected / encrypted OOXML workbooks are not supported")
    try:
        workbook_values = load_workbook(file_path, data_only=True, keep_vba=file_path.suffix.lower() == ".xlsm")
        workbook_formulas = load_workbook(file_path, data_only=False, keep_vba=file_path.suffix.lower() == ".xlsm")
    except Exception as exc:
        raise ValueError(f"Could not read workbook: {exc}") from exc

    sheets_info = {
        "all": [sheet.title for sheet in workbook_values.worksheets],
        "count": len(workbook_values.worksheets),
        "empty": [],
        "hidden": [],
        "very_hidden": [],
    }
    merged_cells = {}
    formula_cells = {}
    formula_errors = {}
    formula_cache_misses = {}
    mixed_types = {}
    empty_rows = {}
    empty_columns = {}
    duplicate_headers_report = {}
    whitespace_headers_report = {}
    structural_rows_report = {}
    notes_rows_report = {}
    date_formats = {}
    single_value_columns = {}
    high_null_columns = {}
    header_bands = {}
    metadata_rows = {}
    empty_edge_columns = {}
    sheet_summaries = {}

    for values_sheet in workbook_values.worksheets:
        sheet_name = values_sheet.title
        formula_sheet = workbook_formulas[sheet_name]

        sheet_empty = is_sheet_empty(values_sheet)
        if sheet_empty:
            sheets_info["empty"].append(sheet_name)

        if values_sheet.sheet_state == "veryHidden":
            sheets_info["very_hidden"].append({"name": sheet_name, "state": values_sheet.sheet_state})
        elif values_sheet.sheet_state != "visible":
            sheets_info["hidden"].append({"name": sheet_name, "state": values_sheet.sheet_state})

        if sheet_name in INTERNAL_SHEETS:
            continue

        plan = detect_header_band(values_sheet)
        data_start_row = min(values_sheet.max_row + 1, plan["header_row"] + 1)
        if len(plan["header_band_rows"]) > 1:
            header_bands[sheet_name] = {"rows": plan["header_band_rows"], "header_row": plan["header_row"]}
        if plan["metadata_rows"]:
            metadata_rows[sheet_name] = plan["metadata_rows"]

        merged = [str(rng) for rng in formula_sheet.merged_cells.ranges]
        if merged:
            merged_cells[sheet_name] = merged

        formulas = scan_formula_cells(formula_sheet)
        if formulas:
            formula_cells[sheet_name] = formulas

        errors = scan_formula_errors(values_sheet)
        if errors:
            formula_errors[sheet_name] = errors

        cache_misses = scan_formula_cache_misses(formula_sheet, values_sheet)
        if cache_misses:
            formula_cache_misses[sheet_name] = cache_misses

        if sheet_empty:
            sheet_summaries[sheet_name] = {"risk": "empty", "issues": ["empty_sheet"]}
            continue

        headers = headers_for_sheet(values_sheet, header_row=plan["header_row"])
        whitespace_headers = header_whitespace(headers)
        if whitespace_headers:
            whitespace_headers_report[sheet_name] = whitespace_headers
        duplicates = duplicate_headers(headers)
        if duplicates:
            duplicate_headers_report[sheet_name] = duplicates
        blank_rows = scan_empty_rows(values_sheet, data_start_row)
        if blank_rows:
            empty_rows[sheet_name] = {"count": len(blank_rows), "rows": blank_rows}
        structural_rows = scan_structural_rows(values_sheet, data_start_row)
        if structural_rows:
            structural_rows_report[sheet_name] = structural_rows
        notes_rows = scan_notes_rows(values_sheet, data_start_row)
        if notes_rows:
            notes_rows_report[sheet_name] = notes_rows
        edge_cols = scan_empty_edge_columns(values_sheet, plan["header_row"])
        if edge_cols["leading"] or edge_cols["trailing"]:
            empty_edge_columns[sheet_name] = edge_cols
        mix, empties, singles, dates, high_nulls = scan_columns(values_sheet, headers, data_start_row)
        if mix:
            mixed_types[sheet_name] = mix
        if empties:
            empty_columns[sheet_name] = empties
        if singles:
            single_value_columns[sheet_name] = singles
        if dates:
            date_formats[sheet_name] = dates
        if high_nulls:
            high_null_columns[sheet_name] = high_nulls

        issue_flags = []
        for name, payload in [
            ("merged_ranges", merged),
            ("formula_errors", errors),
            ("formula_cache_misses", cache_misses),
            ("mixed_types", mix),
            ("duplicate_headers", duplicates),
            ("header_band", header_bands.get(sheet_name)),
            ("metadata_rows", metadata_rows.get(sheet_name)),
            ("empty_rows", blank_rows),
            ("notes_rows", notes_rows),
            ("structural_rows", structural_rows),
            ("empty_columns", empties),
            ("empty_edge_columns", edge_cols if edge_cols["leading"] or edge_cols["trailing"] else None),
        ]:
            if payload:
                issue_flags.append(name)
        risk = "critical" if errors or duplicates or mix else "warning" if issue_flags else "healthy"
        sheet_summaries[sheet_name] = {"risk": risk, "issues": issue_flags, "header_row": plan["header_row"], "data_start_row": data_start_row}

    contract = build_contract("excel_doctor.diagnose")
    report = {
        "contract": contract,
        "schema_version": contract["version"],
        "tool_version": TOOL_VERSION,
        "file": file_path.name,
        "file_type": file_path.suffix.lower(),
        "sheets": sheets_info,
        "merged_cells": merged_cells,
        "formula_cells": formula_cells,
        "formula_errors": formula_errors,
        "formula_cache_misses": formula_cache_misses,
        "mixed_types": mixed_types,
        "empty_rows": empty_rows,
        "empty_columns": empty_columns,
        "duplicate_headers": duplicate_headers_report,
        "whitespace_headers": whitespace_headers_report,
        "structural_rows": structural_rows_report,
        "notes_rows": notes_rows_report,
        "date_formats": date_formats,
        "single_value_columns": single_value_columns,
        "high_null_columns": high_null_columns,
        "header_bands": header_bands,
        "metadata_rows": metadata_rows,
        "empty_edge_columns": empty_edge_columns,
        "sheet_summaries": sheet_summaries,
        "workbook_mode": "workbook-native",
        "limitations": [
            ".xls legacy workbooks are not supported by excel-doctor",
            "Password-protected / encrypted workbooks are rejected rather than repaired",
            "Formula cache misses can be detected, but cached results cannot be reconstructed safely",
        ],
    }
    report["summary"] = build_summary(report)
    report["issue_counts"] = count_issue_events(report)
    report["manual_review_warnings"] = build_manual_review_warnings(report)
    report["workbook_summary"] = {
        "sheets_total": sheets_info["count"],
        "data_sheets_evaluated": len(sheet_summaries),
        "hidden_sheet_count": len(sheets_info["hidden"]),
        "very_hidden_sheet_count": len(sheets_info["very_hidden"]),
        "mode": "workbook-native",
    }
    report["run_summary"] = build_run_summary(
        tool="excel-doctor",
        script="diagnose.py",
        input_path=file_path,
        metrics={
            "sheets_total": sheets_info["count"],
            "issues_found": report["summary"]["issue_count"],
            "issue_categories_triggered": report["summary"]["issue_categories_triggered"],
            "manual_review_warnings": len(report["manual_review_warnings"]),
            "verdict": report["summary"]["verdict"],
            "mode": "workbook-native",
        },
    )
    return report


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "No file path provided. Usage: diagnose.py <file.xlsx|file.xlsm>"}), file=sys.stdout)
        sys.exit(1)
    file_path = Path(sys.argv[1])
    if not file_path.exists():
        print(json.dumps({"error": f"File not found: {file_path}"}), file=sys.stdout)
        sys.exit(1)
    suffix = file_path.suffix.lower()
    if suffix == ".xls":
        print(json.dumps({"error": ".xls is not supported by excel-doctor. Use csv-doctor tabular rescue or convert to .xlsx first."}), file=sys.stdout)
        sys.exit(1)
    if suffix not in (".xlsx", ".xlsm"):
        print(json.dumps({"error": f"Expected an .xlsx/.xlsm file, got: {file_path.suffix}"}), file=sys.stdout)
        sys.exit(1)
    try:
        report = build_report(file_path)
    except Exception as exc:
        print(json.dumps({"error": str(exc)}), file=sys.stdout)
        sys.exit(1)
    print(json.dumps(report, indent=2, ensure_ascii=False))
    sys.exit(0)


if __name__ == "__main__":
    main()
