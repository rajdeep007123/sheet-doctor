#!/usr/bin/env python3
"""
heal.py — CSV Doctor Healer  (v2)

Reads a messy CSV and writes a 3-sheet Excel workbook:

  Sheet 1 — "Clean Data"    rows that were fixed and are ready to use
  Sheet 2 — "Quarantine"    rows that could not be fixed or are unusable
  Sheet 3 — "Change Log"    one entry per individual change made

Usage:
    python skills/csv-doctor/scripts/heal.py [input.csv] [output.xlsx] [--sheet NAME | --all-sheets]

Exit codes:
    0 — completed
    1 — input file not found or unreadable
"""

from __future__ import annotations

import argparse
import csv
import io
import re
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pandas as pd

# loader.py lives in the same directory as this script.
sys.path.insert(0, str(Path(__file__).parent))
from column_detector import analyse_dataframe
from loader import load_file

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed — run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── Paths ─────────────────────────────────────────────────────────────────
HERE  = Path(__file__).parent
ROOT  = HERE.parent.parent.parent
INPUT  = ROOT / "sample-data" / "extreme_mess.csv"
OUTPUT = ROOT / "sample-data" / "extreme_mess_healed.xlsx"

# ── Schema ────────────────────────────────────────────────────────────────
HEADERS = ["Employee Name", "Department", "Date", "Amount", "Currency",
           "Category", "Status", "Notes"]
N_COLS  = len(HEADERS)
COL     = {name: i for i, name in enumerate(HEADERS)}

# Sparse-row thresholds — rows with fewer filled fields than this fraction are quarantined.
# Schema-specific: 50% — we know exactly what 8 fields to expect, so <4 filled is clearly broken.
# Generic: 25% — unknown schema means we quarantine only obviously empty rows (>75% blank).
SPARSE_THRESHOLD_SCHEMA  = 0.50
SPARSE_THRESHOLD_GENERIC = 0.25


# ══════════════════════════════════════════════════════════════════════════
# DATA CLASSES
# ══════════════════════════════════════════════════════════════════════════

@dataclass
class Change:
    original_row_number: int
    column_affected:     str
    original_value:      str
    new_value:           str
    action_taken:        str   # Fixed | Quarantined | Flagged | Removed
    reason:              str

@dataclass
class CleanRow:
    row:          list
    row_num:      int
    was_modified: bool
    needs_review: bool

@dataclass
class QuarantineRow:
    row:    list
    row_num: int
    reason: str


@dataclass
class SemanticPlan:
    enabled: bool
    roles_by_index: dict[int, str]
    confidence_by_index: dict[int, float]
    label_idx: int
    amount_idx: int | None
    currency_idx: int | None
    date_idx: int | None
    fill_down_indices: list[int]


FORMULA_RE = re.compile(r"^\s*=")
TOTAL_LABEL_RE = re.compile(r"\b(grand\s+total|subtotal|sub-total|total|sum)\b", re.IGNORECASE)
NOTES_ROW_RE = re.compile(r"\b(approved|manager|note|comment|memo|generated|report|expense|expenses)\b", re.IGNORECASE)
CURRENCY_SYMBOL_MAP = {"$": "USD", "€": "EUR", "£": "GBP", "₹": "INR", "¥": "JPY"}
STATUS_VALUE_HINTS = set(STATUS_MAP.keys()) if "STATUS_MAP" in globals() else {
    "approved", "approve", "rejected", "reject", "pending", "pending review",
}
ROLE_HEADER_HINTS = {
    "name": ("name", "employee", "emp", "person", "contact"),
    "date": ("date", "dated", "txn date", "transaction", "invoice date", "posted"),
    "amount": ("amount", "cost", "price", "value", "expense", "spend", "salary", "pay", "total"),
    "currency": ("currency", "curr", "fx", "ccy"),
    "status": ("status", "state", "approval", "approved", "decision"),
    "department": ("department", "dept", "division", "team", "unit", "function"),
    "category": ("category", "type", "class", "group", "bucket", "expense type"),
    "notes": ("notes", "note", "comment", "comments", "description", "details", "memo", "remarks"),
}


def _strip_nulls(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\x00", "")


def _non_empty_cells(row: list[str]) -> list[str]:
    cells = []
    for cell in row:
        cleaned = _strip_nulls(cell).strip()
        if cleaned:
            cells.append(cleaned)
    return cells


def _joined_row_text(row: list[str]) -> str:
    return " | ".join(_strip_nulls(cell).strip() for cell in row if _strip_nulls(cell).strip())


def _looks_like_header_row(row: list[str]) -> bool:
    if is_schema_specific_header(row):
        return True

    non_empty = [_strip_nulls(cell).strip() for cell in row if _strip_nulls(cell).strip()]
    if len(non_empty) < 2:
        return False
    if any(FORMULA_RE.match(cell) for cell in non_empty):
        return False
    if sum(1 for cell in non_empty if len(cell) > 50) >= 2:
        return False
    data_like_cells = 0
    for cell in non_empty:
        lower_cell = cell.strip().lower()
        if parse_amount_like(cell) is not None:
            data_like_cells += 1
            continue
        extracted_amount, _ = extract_currency_from_text(cell)
        if extracted_amount and parse_amount_like(extracted_amount) is not None:
            data_like_cells += 1
            continue
        normalized_date, changed, _ = normalise_date(cell)
        if changed or re.match(r"^\d{4}-\d{2}-\d{2}$", normalized_date):
            data_like_cells += 1
            continue
        if lower_cell in STATUS_VALUE_HINTS:
            data_like_cells += 1
    if data_like_cells >= 2:
        return False
    alpha_cells = sum(1 for cell in non_empty if re.search(r"[A-Za-z]", cell))
    numeric_heavy = sum(1 for cell in non_empty if re.fullmatch(r"[\d,./:-]+", cell))
    return alpha_cells >= max(2, len(non_empty) - 1) and numeric_heavy <= 1


def detect_header_row_index(all_rows: list[list[str]]) -> int:
    search_rows = all_rows[:10]
    exact_matches = [
        idx for idx, row in enumerate(search_rows)
        if is_schema_specific_header(row) and idx < len(all_rows) - 1
    ]
    if exact_matches:
        return exact_matches[-1]

    generic_candidates = [
        idx for idx, row in enumerate(search_rows)
        if idx < len(all_rows) - 1 and _looks_like_header_row(row)
    ]
    if generic_candidates:
        return generic_candidates[-1]
    return 0


def preprocess_rows(all_rows: list[list[str]]) -> tuple[list[list[str]], list[Change]]:
    header_idx = detect_header_row_index(all_rows)
    if header_idx <= 0:
        return all_rows, []

    changes: list[Change] = []
    for i, row in enumerate(all_rows[:header_idx], start=1):
        row_text = _joined_row_text(row) or "[empty metadata row]"
        changes.append(
            Change(
                i,
                "[file metadata]",
                row_text[:200],
                "",
                "Removed",
                "File Metadata: row before detected header moved out of the dataset",
            )
        )

    return all_rows[header_idx:], changes


def is_formula_residue(value: str) -> bool:
    return bool(FORMULA_RE.match(_strip_nulls(value).strip()))


def detect_formula_row(row: list[str], headers: list[str]) -> tuple[bool, str]:
    for idx, cell in enumerate(row):
        if is_formula_residue(cell):
            label = headers[idx] if idx < len(headers) else f"[col {idx + 1}]"
            return True, label
    return False, ""


def looks_like_notes_row(row: list[str]) -> bool:
    non_empty = _non_empty_cells(row)
    if len(non_empty) != 1:
        return False
    text = non_empty[0]
    if len(text) <= 50:
        return False
    if len(text.split()) < 8:
        return False
    return bool(NOTES_ROW_RE.search(text) or re.search(r"[A-Za-z]{4,}", text))


def parse_amount_like(value: str) -> float | None:
    if not value.strip():
        return None
    normalised, changed, _ = normalise_amount(value)
    candidate = normalised if changed or normalised != value else value.strip()
    try:
        return float(candidate)
    except ValueError:
        return None


def row_amount_totalish(label_cell: str, amount_cell: str, running_total: float) -> bool:
    if not TOTAL_LABEL_RE.search(label_cell or ""):
        return False
    amount = parse_amount_like(amount_cell or "")
    if amount is None:
        return False
    tolerance = max(1.0, abs(running_total) * 0.02)
    return abs(amount - running_total) <= tolerance


def sparse_total_label_row(row: list[str], label_idx: int, amount_idx: int) -> bool:
    label_cell = row[label_idx] if label_idx < len(row) else ""
    amount_cell = row[amount_idx] if amount_idx < len(row) else ""
    if not TOTAL_LABEL_RE.search(label_cell or ""):
        return False
    if parse_amount_like(amount_cell or "") is None:
        return False
    non_empty = sum(1 for cell in row if cell.strip())
    return non_empty <= 2


def extract_currency_from_text(value: str) -> tuple[str | None, str | None]:
    raw = value.strip()
    if not raw:
        return None, None

    code_match = re.search(r"\b(USD|EUR|GBP|INR|CAD|AUD|JPY)\b", raw, flags=re.IGNORECASE)
    symbol_match = next((symbol for symbol in CURRENCY_SYMBOL_MAP if symbol in raw), None)
    currency = None
    if code_match:
        currency = code_match.group(1).upper()
    elif symbol_match:
        currency = CURRENCY_SYMBOL_MAP[symbol_match]

    amount_candidate = raw
    if code_match:
        amount_candidate = re.sub(r"\b(USD|EUR|GBP|INR|CAD|AUD|JPY)\b", "", amount_candidate, flags=re.IGNORECASE)
    if symbol_match:
        amount_candidate = amount_candidate.replace(symbol_match, "")
    amount_candidate = " ".join(amount_candidate.split()).strip()

    if currency and parse_amount_like(amount_candidate) is not None:
        return amount_candidate, currency
    return None, currency



# ══════════════════════════════════════════════════════════════════════════
# STEP 1 — Read with mixed-encoding tolerance (via loader)
# ══════════════════════════════════════════════════════════════════════════

def read_file(
    path: Path,
    *,
    sheet_name: str | None = None,
    consolidate_sheets: bool | None = None,
) -> tuple[list[list[str]], str]:
    """
    Load any supported file format via loader.load_file().

    Returns (raw_rows, delimiter).  For non-text formats (Excel, ODS, JSON)
    raw_rows are reconstructed from the DataFrame so the rest of the
    processing pipeline stays unchanged.
    """
    result    = load_file(path, sheet_name=sheet_name, consolidate_sheets=consolidate_sheets)
    raw_text  = result["raw_text"]
    delimiter = result["delimiter"]

    if raw_text is not None and delimiter is not None:
        # Text format: re-parse with csv.reader so multi-line quoted fields
        # are reconstructed correctly (same behaviour as the old read_mixed_encoding).
        rows = list(csv.reader(io.StringIO(raw_text), delimiter=delimiter))
    else:
        # Binary/structured format (Excel, ODS, JSON…): convert the DataFrame
        # back to rows so the rest of the pipeline is format-agnostic.
        df        = result["dataframe"]
        delimiter = ","
        rows      = [list(df.columns)] + [
            [str(v) if v is not None else "" for v in row]
            for row in df.itertuples(index=False, name=None)
        ]

    return rows, delimiter or ","


# ══════════════════════════════════════════════════════════════════════════
# STEP 2 — Row classification
# ══════════════════════════════════════════════════════════════════════════

QUARANTINE_REASONS = {
    "EMPTY":               "Completely empty row",
    "WHITESPACE":          "Row is all whitespace",
    "STRUCTURAL_HEADER":   "Structural row (TOTAL/subtotal/header repeat)",
    "STRUCTURAL_TOTAL":    "Structural row (TOTAL/subtotal/header repeat)",
    "CALCULATED_SUBTOTAL": "Calculated subtotal row",
    "NOTES_ROW":           "Appears to be a notes row",
    "FORMULA":             "Excel formula found, not data",
    "SPARSE":              f"Less than {int(SPARSE_THRESHOLD_SCHEMA * 100)}% columns filled",
}

def classify_raw_row(row: list[str], header_sig: tuple[str, ...]) -> str:
    stripped = [c.strip() for c in row]

    if all(c == "" for c in stripped):
        # Distinguish truly empty (all "") from whitespace-only
        return "WHITESPACE" if any(c != "" for c in row) else "EMPTY"

    if tuple(c.lower() for c in stripped) == header_sig:
        return "STRUCTURAL_HEADER"

    if looks_like_notes_row(row):
        return "NOTES_ROW"

    if any(is_formula_residue(cell) for cell in stripped):
        return "FORMULA"

    if TOTAL_LABEL_RE.search(stripped[0]) and len(stripped) > COL["Amount"]:
        if parse_amount_like(stripped[COL["Amount"]]) is not None:
            return "NORMAL"

    non_empty = sum(1 for c in stripped if c)
    if non_empty < N_COLS * SPARSE_THRESHOLD_SCHEMA:
        return "SPARSE"

    return "NORMAL"


# ══════════════════════════════════════════════════════════════════════════
# STEP 3 — Alignment fix (with Change log entry)
# ══════════════════════════════════════════════════════════════════════════

def fix_alignment(row: list[str], row_num: int) -> tuple[list[str], Change | None]:
    """Detect and fix three structural column problems."""
    n = len(row)
    if n == N_COLS:
        return row, None

    if n > N_COLS:
        # Shifted right: leading empty ghost column
        if row[0].strip() == "" and row[1].strip() != "":
            fixed = row[1: N_COLS + 1]
            return fixed, Change(
                row_num, "[row structure]",
                f"{n} columns (empty leading ghost col)",
                f"{N_COLS} columns",
                "Fixed", "Shifted-right row: empty leading column stripped"
            )
        # Phantom comma: empty ghost field sits between Status and Notes
        if n == N_COLS + 1 and row[N_COLS - 1].strip() == "" and row[N_COLS].strip() != "":
            fixed = row[: N_COLS - 1] + [row[N_COLS]]
            return fixed, Change(
                row_num, "Notes",
                f"[ghost field] + '{row[N_COLS]}'",
                row[N_COLS],
                "Fixed", "Phantom comma: empty ghost field before Notes removed"
            )
        # Unquoted commas in Notes: merge overflow columns back
        merged = ", ".join(row[N_COLS - 1:]).strip()
        fixed = row[: N_COLS - 1] + [merged]
        return fixed, Change(
            row_num, "Notes",
            f"{n - N_COLS + 1} fragments: {row[N_COLS - 1]!r}...",
            merged,
            "Fixed", "Unquoted commas in Notes field: fragments merged into one"
        )

    # Short row: pad with empty strings
    fixed = row + [""] * (N_COLS - n)
    return fixed, Change(
        row_num, "[row structure]",
        f"{n} columns",
        f"{N_COLS} columns ({N_COLS - n} empty field(s) appended)",
        "Fixed", f"Short row padded with {N_COLS - n} empty field(s)"
    )


# ══════════════════════════════════════════════════════════════════════════
# STEP 4 — Cell-level cleaning (BOM, null, line breaks, smart quotes)
# ══════════════════════════════════════════════════════════════════════════

SMART_QUOTES = {
    "\u201c": '"', "\u201d": '"',   # "" double curly
    "\u2018": "'", "\u2019": "'",   # '' single curly / smart apostrophe
}

def _clean_cell_text(value: object) -> tuple[str, list[str]]:
    """Strip BOM, null bytes, line breaks, smart quotes. Returns (cleaned_value, reasons)."""
    new_val = _strip_nulls(value)
    reasons = []

    if "\ufeff" in new_val:
        new_val = new_val.replace("\ufeff", "")
        reasons.append("BOM byte-order mark stripped")
    if "\x00" in new_val:
        new_val = new_val.replace("\x00", "")
        reasons.append("Null byte removed")
    if "\n" in new_val or "\r" in new_val:
        new_val = new_val.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
        reasons.append("Embedded line break replaced with space")

    had_smart = any(s in new_val for s in SMART_QUOTES)
    for smart, straight in SMART_QUOTES.items():
        new_val = new_val.replace(smart, straight)
    if had_smart:
        reasons.append("Smart/curly quotes normalised to straight quotes")

    new_val = new_val.strip()
    return new_val, reasons


def _col_name(i: int) -> str:
    return HEADERS[i] if i < N_COLS else f"[col {i + 1}]"

def clean_row(row: list[str], row_num: int) -> tuple[list[str], list[Change]]:
    cleaned, changes = [], []
    for i, cell in enumerate(row):
        new_val, reasons = _clean_cell_text(cell)
        if reasons:
            orig_display = (cell
                            .replace("\ufeff", "[BOM]")
                            .replace("\x00", "[NULL]")
                            .strip())
            changes.append(Change(
                row_num, _col_name(i),
                orig_display, new_val,
                "Fixed", "; ".join(reasons)
            ))
        cleaned.append(new_val)
    return cleaned, changes


# ══════════════════════════════════════════════════════════════════════════
# STEP 5 — Value normalisation
# ══════════════════════════════════════════════════════════════════════════

EXCEL_EPOCH  = datetime(1899, 12, 30)
MONTH_NAMES  = {
    "january":1,"february":2,"march":3,"april":4,"may":5,"june":6,
    "july":7,"august":8,"september":9,"october":10,"november":11,"december":12,
    "jan":1,"feb":2,"mar":3,"apr":4,"jun":6,"jul":7,"aug":8,
    "sep":9,"oct":10,"nov":11,"dec":12,
}

def _fmt(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d")

def normalise_date(value: str) -> tuple[str, bool, str]:
    v = value.strip()
    if not v or re.match(r"^\d{4}-\d{2}-\d{2}$", v):
        return v, False, ""

    # ISO 8601 with time: 2023-01-18T00:00:00Z
    m = re.match(r"^(\d{4}-\d{2}-\d{2})T", v)
    if m:
        return m.group(1), True, "ISO 8601 datetime truncated to date-only"

    # DD/MM/YYYY or MM/DD/YYYY — try day-first, fall back to month-first
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", v)
    if m:
        a, b, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        for day, month, fmt in [(a, b, "DD/MM/YYYY"), (b, a, "MM/DD/YYYY")]:
            try:
                return _fmt(datetime(year, month, day)), True, \
                       f"{fmt} normalised to ISO YYYY-MM-DD (day-first assumed for ambiguous dates)"
            except ValueError:
                continue
        return v, False, ""

    # MM-DD-YY (two-digit year, hyphens)
    m = re.match(r"^(\d{2})-(\d{2})-(\d{2})$", v)
    if m:
        mo, day, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        year = 2000 + yr if yr < 50 else 1900 + yr
        try:
            return _fmt(datetime(year, mo, day)), True, \
                   f"MM-DD-YY normalised (20xx assumed for year < 50)"
        except ValueError:
            return v, False, ""

    # Month DD YYYY or Month D YYYY
    m = re.match(r"^([A-Za-z]+)\s+(\d{1,2})\s+(\d{4})$", v)
    if m:
        month_num = MONTH_NAMES.get(m.group(1).lower())
        if month_num:
            try:
                return _fmt(datetime(int(m.group(3)), month_num, int(m.group(2)))), True, \
                       "Written-out month name normalised to ISO YYYY-MM-DD"
            except ValueError:
                pass
        return v, False, ""

    # Unix timestamp (10 digits)
    if re.match(r"^\d{10}$", v):
        try:
            dt = datetime.fromtimestamp(int(v), tz=timezone.utc)
            return _fmt(dt), True, "Unix timestamp (UTC) converted to YYYY-MM-DD"
        except (ValueError, OSError):
            return v, False, ""

    # Excel serial date (5-digit integer in plausible range)
    if re.match(r"^\d{5}$", v) and 40_000 <= int(v) <= 55_000:
        return _fmt(EXCEL_EPOCH + timedelta(days=int(v))), True, \
               "Excel serial date (Windows epoch 1899-12-30) converted to YYYY-MM-DD"

    return v, False, ""


_AMOUNT_NULL = {"n/a", "tbd", "-", "na", "nil", "none", ""}

def normalise_amount(value: str) -> tuple[str, bool, str]:
    v = value.strip()
    orig = v
    if v.lower() in _AMOUNT_NULL:
        result = ""
        return result, (result != orig), "Non-numeric placeholder left blank (N/A / TBD)"

    # Strip currency symbols and trailing ISO codes
    v = re.sub(r"[€£¥₹$]", "", v)
    v = re.sub(r"\s*(USD|EUR|GBP|INR|CAD|AUD)\s*$", "", v, flags=re.IGNORECASE).strip()

    # Negative accounting notation: (500) → -500
    m = re.match(r"^\(([0-9,. ]+)\)$", v)
    if m:
        v = "-" + m.group(1)

    # European decimal: 1.200,00  (period = thousands, comma = decimal)
    if "," in v and "." in v:
        if v.index(".") < v.index(","):
            v = v.replace(".", "").replace(",", ".")
            desc = "European decimal format (1.200,00) converted"
        else:
            v = v.replace(",", "")
            desc = "US thousands separator removed"
    elif "," in v:
        if re.search(r",\d{2}$", v):
            v = v.replace(",", ".")
            desc = "Comma decimal separator converted to period"
        else:
            v = v.replace(",", "")
            desc = "Thousands-separator comma removed"
    else:
        desc = "Amount normalised to 2 decimal places"

    try:
        result = f"{float(v):.2f}"
        return result, (result != orig), desc
    except ValueError:
        return orig, False, ""


CURRENCY_MAP = {
    "usd": "USD", "us dollar": "USD", "u.s. dollar": "USD", "dollar": "USD", "$": "USD",
    "eur": "EUR", "euro": "EUR", "€": "EUR",
    "gbp": "GBP", "pound": "GBP", "sterling": "GBP", "£": "GBP",
    "inr": "INR", "rupee": "INR", "indian rupee": "INR", "₹": "INR",
    "cad": "CAD", "canadian dollar": "CAD",
    "aud": "AUD", "australian dollar": "AUD",
}

def normalise_currency(value: str) -> tuple[str, bool, str]:
    v = value.strip()
    if not v:
        return v, False, ""
    cleaned = v.replace("₹", "").replace("€", "").replace("$", "").replace("£", "").strip()
    for lookup in (cleaned.lower(), v.lower()):
        if lookup in CURRENCY_MAP:
            result = CURRENCY_MAP[lookup]
            return result, result != v, f"Currency '{v}' standardised to ISO 3-letter code"
    if re.match(r"^[A-Z]{3}$", cleaned):
        return cleaned, cleaned != v, "Currency uppercased to ISO format"
    return v, False, ""


def split_amount_currency_fields(row: list[str], row_num: int) -> tuple[list[str], list[Change]]:
    return split_amount_currency_fields_dynamic(
        row,
        row_num,
        HEADERS,
        COL["Amount"],
        COL["Currency"],
    )


def normalise_name(value: str) -> tuple[str, bool, str]:
    v = " ".join(value.split())   # collapse multiple spaces
    if not v:
        return v, False, ""
    if "," in v:
        parts = [p.strip() for p in v.split(",", 1)]
        if len(parts) == 2 and parts[1]:
            v = f"{parts[1]} {parts[0]}"
    result = v.title()
    reasons = []
    if " ".join(value.split()) != value:
        reasons.append("extra whitespace collapsed")
    if "," in value:
        reasons.append("Last, First → First Last")
    if result != " ".join(value.split()).title():
        pass
    if result != value:
        if not reasons:
            reasons.append("name title-cased")
    return result, result != value, "Name normalised: " + "; ".join(reasons) if reasons else "Name title-cased"


STATUS_MAP = {
    "approved":       "Approved",
    "approve":        "Approved",
    "rejected":       "Rejected",
    "reject":         "Rejected",
    "pending":        "Pending",
    "pending review": "Pending",
}

def normalise_status(value: str) -> tuple[str, bool, str]:
    v = value.strip()
    result = STATUS_MAP.get(v.lower(), v.title() if v else v)
    reason  = f"Status '{v}' standardised to canonical form" if result != v else ""
    return result, result != v, reason


def apply_normalisations(row: list[str], row_num: int) -> tuple[list[str], list[Change]]:
    row     = row.copy()
    changes = []

    row, split_changes = split_amount_currency_fields(row, row_num)
    changes.extend(split_changes)

    def maybe_fix(idx, fn, col_label):
        orig = row[idx]
        new, changed, reason = fn(orig)
        if changed:
            row[idx] = new
            changes.append(Change(row_num, col_label, orig, new, "Fixed", reason))

    maybe_fix(COL["Date"],          normalise_date,     "Date")
    maybe_fix(COL["Amount"],        normalise_amount,   "Amount")
    maybe_fix(COL["Currency"],      normalise_currency, "Currency")
    maybe_fix(COL["Employee Name"], normalise_name,     "Employee Name")
    maybe_fix(COL["Status"],        normalise_status,   "Status")

    # Department — title-case + collapse whitespace
    orig = row[COL["Department"]]
    fixed_dept = " ".join(orig.split()).title()
    if fixed_dept != orig:
        row[COL["Department"]] = fixed_dept
        changes.append(Change(row_num, "Department", orig, fixed_dept,
                               "Fixed", "Department title-cased"))

    # Category — title-case
    orig = row[COL["Category"]]
    fixed_cat = orig.strip().title()
    if fixed_cat != orig:
        row[COL["Category"]] = fixed_cat
        changes.append(Change(row_num, "Category", orig, fixed_cat,
                               "Fixed", "Category title-cased"))

    return row, changes


# ══════════════════════════════════════════════════════════════════════════
# STEP 6 — needs_review decision
# ══════════════════════════════════════════════════════════════════════════

def needs_review(row: list[str], was_padded: bool) -> bool:
    """Row needs human review if: amount blank, date unparseable, or padded."""
    if not row[COL["Amount"]]:
        return True
    date_val = row[COL["Date"]]
    if date_val and not re.match(r"^\d{4}-\d{2}-\d{2}$", date_val):
        return True
    return was_padded


def forward_fill_merged_cell_gaps_generic(
    clean_data: list[CleanRow], changelog: list[Change], headers: list[str], fill_columns: list[int]
) -> None:
    for col_idx in fill_columns:
        last_value = ""
        gap_indexes: list[int] = []

        for idx, entry in enumerate(clean_data):
            cell_value = entry.row[col_idx].strip()
            other_filled = sum(1 for j, cell in enumerate(entry.row) if j != col_idx and cell.strip())

            if cell_value:
                if last_value and gap_indexes and len(gap_indexes) <= 5:
                    for gap_idx in gap_indexes:
                        gap_entry = clean_data[gap_idx]
                        if gap_entry.row[col_idx].strip():
                            continue
                        gap_entry.row[col_idx] = last_value
                        gap_entry.was_modified = True
                        gap_entry.needs_review = True
                        changelog.append(
                            Change(
                                gap_entry.row_num,
                                headers[col_idx],
                                "",
                                last_value,
                                "Fixed",
                                "Blank categorical cell forward-filled to repair a merged-cell style export gap",
                            )
                        )
                last_value = cell_value
                gap_indexes = []
                continue

            if last_value and other_filled >= 2:
                gap_indexes.append(idx)
            else:
                gap_indexes = []


def forward_fill_merged_cell_gaps(clean_data: list[CleanRow], changelog: list[Change]) -> None:
    forward_fill_merged_cell_gaps_generic(
        clean_data,
        changelog,
        HEADERS,
        [COL["Department"], COL["Category"], COL["Status"], COL["Currency"]],
    )


def flag_near_duplicates_semantic(
    clean_data: list[CleanRow],
    changelog: list[Change],
    headers: list[str],
    semantic_plan: SemanticPlan,
) -> None:
    if semantic_plan.date_idx is None or semantic_plan.amount_idx is None:
        return

    key_indices = [
        idx for idx, role in semantic_plan.roles_by_index.items()
        if role in {"name", "amount", "currency", "category", "department"}
    ]
    if len(key_indices) < 2:
        return

    nd_index: dict[tuple[str, ...], int] = {}
    for idx, entry in enumerate(clean_data):
        row = entry.row
        key = tuple(row[i] for i in key_indices)
        if key in nd_index:
            prev = clean_data[nd_index[key]]
            d1 = prev.row[semantic_plan.date_idx]
            d2 = entry.row[semantic_plan.date_idx]
            if (
                d1
                and d2
                and re.match(r"^\d{4}-\d{2}-\d{2}$", d1)
                and re.match(r"^\d{4}-\d{2}-\d{2}$", d2)
            ):
                try:
                    delta = abs(
                        (
                            datetime.strptime(d2, "%Y-%m-%d")
                            - datetime.strptime(d1, "%Y-%m-%d")
                        ).days
                    )
                except ValueError:
                    continue
                if delta <= 2:
                    prev.needs_review = True
                    entry.needs_review = True
                    label_idx = semantic_plan.label_idx
                    for flagged, other_date, other_row_num in [
                        (entry, d1, prev.row_num),
                        (prev, d2, entry.row_num),
                    ]:
                        changelog.append(
                            Change(
                                flagged.row_num,
                                "[row]",
                                flagged.row[label_idx] if label_idx < len(flagged.row) else "",
                                "",
                                "Flagged",
                                f"Near-duplicate: same semantic key columns; date {flagged.row[semantic_plan.date_idx]} differs by {delta} day(s) from row {other_row_num} ({other_date})",
                            )
                        )
        else:
            nd_index[key] = idx


# ══════════════════════════════════════════════════════════════════════════
# MAIN PROCESSING LOOP
# ══════════════════════════════════════════════════════════════════════════

def process_schema_specific(
    all_rows: list[list[str]],
    initial_changelog: list[Change] | None = None,
) -> tuple[list[CleanRow], list[QuarantineRow], list[Change]]:
    header_sig = tuple(c.strip().lower() for c in all_rows[0])

    clean_data: list[CleanRow]       = []
    quarantine: list[QuarantineRow]  = []
    changelog:  list[Change]         = list(initial_changelog or [])
    seen_exact: dict[tuple, int]     = {}   # normalized row → original row_num
    running_amount_total = 0.0

    # Skip row 0 (actual column headers); start from row 1 (metadata / first data row)
    data_rows = all_rows[1:]

    for i, raw_row in enumerate(data_rows):
        row_num = i + 2   # 1-based; header = row 1

        # ── Classify on raw values ────────────────────────────────────────
        cls = classify_raw_row(raw_row, header_sig)

        if cls != "NORMAL":
            q_reason = QUARANTINE_REASONS[cls]
            # Light-clean the row for display in Quarantine tab
            q_row = [c.strip() for c in raw_row]
            q_row = (q_row + [""] * N_COLS)[:N_COLS]          # normalise length
            row_id = next((c for c in q_row if c), "[empty]")
            column_hint = "[row]"
            if cls == "FORMULA":
                _, formula_column = detect_formula_row(raw_row, HEADERS)
                column_hint = formula_column or "formula_residue"
            quarantine.append(QuarantineRow(q_row, row_num, q_reason))
            changelog.append(Change(
                row_num, column_hint, row_id[:60], "",
                "Quarantined",
                "formula_residue: Excel formula found, not data" if cls == "FORMULA" else q_reason
            ))
            continue

        # ── Fix alignment ─────────────────────────────────────────────────
        aligned, align_chg = fix_alignment(raw_row, row_num)
        was_padded = align_chg is not None and "padded" in align_chg.reason.lower()
        if align_chg:
            changelog.append(align_chg)

        # ── Clean cells ───────────────────────────────────────────────────
        cleaned, cell_chgs = clean_row(aligned, row_num)
        changelog.extend(cell_chgs)

        # ── Normalise values ──────────────────────────────────────────────
        fixed, norm_chgs = apply_normalisations(cleaned, row_num)
        changelog.extend(norm_chgs)

        label_cell = fixed[COL["Employee Name"]] or fixed[COL["Department"]] or fixed[COL["Category"]]
        if row_amount_totalish(label_cell, fixed[COL["Amount"]], running_amount_total) or sparse_total_label_row(fixed, COL["Employee Name"], COL["Amount"]):
            quarantine.append(QuarantineRow(fixed, row_num, QUARANTINE_REASONS["CALCULATED_SUBTOTAL"]))
            changelog.append(
                Change(
                    row_num,
                    "Amount",
                    fixed[COL["Amount"]],
                    "",
                    "Quarantined",
                    "Calculated subtotal row",
                )
            )
            continue

        was_modified = bool(align_chg or cell_chgs or norm_chgs)

        # ── Exact-duplicate removal ───────────────────────────────────────
        row_key = tuple(fixed)
        if row_key in seen_exact:
            first_num = seen_exact[row_key]
            changelog.append(Change(
                row_num, "[row]", fixed[0], "",
                "Removed", f"Exact duplicate of row {first_num}"
            ))
            continue
        seen_exact[row_key] = row_num

        clean_data.append(CleanRow(
            row          = fixed,
            row_num      = row_num,
            was_modified = was_modified,
            needs_review = needs_review(fixed, was_padded),
        ))
        parsed_amount = parse_amount_like(fixed[COL["Amount"]])
        if parsed_amount is not None:
            running_amount_total += parsed_amount

    forward_fill_merged_cell_gaps(clean_data, changelog)

    # ── Near-duplicate detection (second pass on clean_data) ─────────────
    nd_index: dict[tuple, int] = {}   # key → index in clean_data
    for idx, entry in enumerate(clean_data):
        r = entry.row
        key = (r[COL["Employee Name"]], r[COL["Amount"]],
               r[COL["Currency"]],     r[COL["Category"]])
        if key in nd_index:
            j       = nd_index[key]
            prev    = clean_data[j]
            d1, d2  = prev.row[COL["Date"]], entry.row[COL["Date"]]
            if (d1 and d2
                    and re.match(r"^\d{4}-\d{2}-\d{2}$", d1)
                    and re.match(r"^\d{4}-\d{2}-\d{2}$", d2)):
                try:
                    delta = abs((datetime.strptime(d2, "%Y-%m-%d")
                                 - datetime.strptime(d1, "%Y-%m-%d")).days)
                    if delta <= 2:
                        prev.needs_review  = True
                        entry.needs_review = True
                        for flagged, other_date, other_row_num in [
                            (entry, d1, prev.row_num),
                            (prev,  d2, entry.row_num),
                        ]:
                            changelog.append(Change(
                                flagged.row_num, "[row]",
                                flagged.row[COL["Employee Name"]], "",
                                "Flagged",
                                f"Near-duplicate: same Name/Amount/Currency/Category; "
                                f"date {flagged.row[COL['Date']]} differs by {delta} day(s) "
                                f"from row {other_row_num} ({other_date})"
                            ))
                except ValueError:
                    pass
        else:
            nd_index[key] = idx

    return clean_data, quarantine, changelog


GENERIC_QUARANTINE_REASONS = {
    "EMPTY":             "Completely empty row",
    "WHITESPACE":        "Row is all whitespace",
    "STRUCTURAL_HEADER": "Structural row (header repeated)",
    "STRUCTURAL_TOTAL":  "Structural row (TOTAL/subtotal)",
    "CALCULATED_SUBTOTAL": "Calculated subtotal row",
    "NOTES_ROW":         "Appears to be a notes row",
    "FORMULA":           "Excel formula found, not data",
    "SPARSE":            f"Less than {int(SPARSE_THRESHOLD_GENERIC * 100)}% columns filled",
}


def _normalise_header_text(value: str, index: int) -> str:
    cleaned = " ".join(value.replace("\ufeff", "").strip().split())
    if not cleaned:
        return f"column_{index}"
    return cleaned


def normalise_headers_generic(raw_header: list[str]) -> tuple[list[str], list[Change]]:
    headers = []
    changes: list[Change] = []
    seen = Counter()

    for i, cell in enumerate(raw_header, start=1):
        original = cell or ""
        cleaned, clean_reasons = _clean_cell_text(original)
        base = _normalise_header_text(cleaned, i)
        key = base.lower()
        seen[key] += 1
        final = f"{base}_{seen[key]}" if seen[key] > 1 else base

        reasons = list(clean_reasons)
        if seen[key] > 1:
            reasons.append("Duplicate header renamed with suffix")
        if final != original:
            reason = "; ".join(reasons) if reasons else "Header normalised"
            changes.append(Change(1, f"[header col {i}]", original, final, "Fixed", reason))

        headers.append(final)

    return headers, changes


def classify_raw_row_generic(row: list[str], header_sig: tuple[str, ...], n_cols: int) -> str:
    stripped = [c.strip() for c in row]

    if all(c == "" for c in stripped):
        return "WHITESPACE" if any(c != "" for c in row) else "EMPTY"

    if tuple(c.lower() for c in stripped) == header_sig:
        return "STRUCTURAL_HEADER"

    if looks_like_notes_row(row):
        return "NOTES_ROW"

    if any(is_formula_residue(cell) for cell in stripped):
        return "FORMULA"

    first_non_empty = next((c for c in stripped if c), "")
    non_empty = sum(1 for c in stripped if c)

    if re.match(r"^(grand\s+total|subtotal|total)\b", first_non_empty, flags=re.IGNORECASE):
        if non_empty <= max(2, n_cols // 3):
            return "STRUCTURAL_TOTAL"

    if non_empty < max(1, int(n_cols * SPARSE_THRESHOLD_GENERIC)):
        return "SPARSE"

    return "NORMAL"


def fix_alignment_generic(
    row: list[str], row_num: int, n_cols: int, delimiter: str
) -> tuple[list[str], Change | None, bool]:
    n = len(row)
    if n == n_cols:
        return row, None, False

    if n > n_cols:
        merged_tail = f"{delimiter} ".join(part for part in row[n_cols - 1:] if part != "")
        fixed = row[: n_cols - 1] + [merged_tail]
        return fixed, Change(
            row_num,
            "[row structure]",
            f"{n} columns",
            f"{n_cols} columns",
            "Fixed",
            f"Overflow columns merged into last column using delimiter '{delimiter}'",
        ), True

    fixed = row + [""] * (n_cols - n)
    return fixed, Change(
        row_num,
        "[row structure]",
        f"{n} columns",
        f"{n_cols} columns ({n_cols - n} empty field(s) appended)",
        "Fixed",
        f"Short row padded with {n_cols - n} empty field(s)",
    ), True


def clean_row_generic(
    row: list[str], row_num: int, headers: list[str]
) -> tuple[list[str], list[Change]]:
    cleaned = []
    changes: list[Change] = []

    for i, cell in enumerate(row):
        original = cell
        new_val, reasons = _clean_cell_text(cell)
        if reasons:
            col_label = headers[i] if i < len(headers) else f"[col {i + 1}]"
            orig_display = original.replace("\ufeff", "[BOM]").replace("\x00", "[NULL]")
            changes.append(
                Change(row_num, col_label, orig_display, new_val, "Fixed", "; ".join(reasons))
            )
        cleaned.append(new_val)

    return cleaned, changes


def _header_text(header: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (header or "").strip().lower()).strip()


def _header_matches_role(header: str, role: str) -> bool:
    lowered = _header_text(header)
    return any(token in lowered for token in ROLE_HEADER_HINTS.get(role, ()))


def _status_like_column(column_stats: dict) -> bool:
    values = [
        (entry.get("value") or "").strip().lower()
        for entry in column_stats.get("most_common_values", [])
        if (entry.get("value") or "").strip()
    ]
    return bool(values) and sum(1 for value in values if value in STATUS_VALUE_HINTS) >= max(1, len(values) // 2)


def _average_sample_length(column_stats: dict) -> float:
    samples = [value for value in column_stats.get("sample_values", []) if value]
    if not samples:
        return 0.0
    return sum(len(value) for value in samples) / len(samples)


def _semantic_role_scores(header: str, column_stats: dict) -> dict[str, float]:
    detected_type = column_stats.get("detected_type", "unknown")
    scores = {
        "name": 0.0,
        "date": 0.0,
        "amount": 0.0,
        "currency": 0.0,
        "status": 0.0,
        "department": 0.0,
        "category": 0.0,
        "notes": 0.0,
    }

    if detected_type == "name":
        scores["name"] += 0.72
    if detected_type == "date":
        scores["date"] += 0.72
    if detected_type == "currency/amount":
        scores["amount"] += 0.72
    if detected_type == "plain number":
        scores["amount"] += 0.42
    if detected_type == "currency code":
        scores["currency"] += 0.72
    if detected_type == "boolean":
        scores["status"] += 0.20
    if detected_type == "categorical":
        scores["status"] += 0.12
        scores["department"] += 0.12
        scores["category"] += 0.12
    if detected_type == "free text":
        scores["notes"] += 0.20

    for role in scores:
        if _header_matches_role(header, role):
            if role in {"name", "currency"}:
                scores[role] += 0.68
            elif role in {"date", "amount"}:
                scores[role] += 0.32
            else:
                scores[role] += 0.68

    if _status_like_column(column_stats):
        scores["status"] += 0.28
    if detected_type == "free text" and _average_sample_length(column_stats) >= 20:
        scores["notes"] += 0.12

    return {role: min(score, 0.99) for role, score in scores.items()}


def build_semantic_plan(headers: list[str], raw_rows: list[list[str]], delimiter: str) -> SemanticPlan:
    n_cols = len(headers)
    preview_rows: list[list[str]] = []

    for idx, raw_row in enumerate(raw_rows[:1000], start=2):
        aligned, _, _ = fix_alignment_generic(raw_row, idx, n_cols, delimiter)
        cleaned, _ = clean_row_generic(aligned, idx, headers)
        preview_rows.append(cleaned)

    if not preview_rows:
        return SemanticPlan(False, {}, {}, 0, None, None, None, [])

    analysis = analyse_dataframe(pd.DataFrame(preview_rows, columns=headers))
    columns = analysis.get("columns", {})
    candidate_scores = {
        index: _semantic_role_scores(header, columns.get(header, {}))
        for index, header in enumerate(headers)
    }

    thresholds = {
        "name": 0.60,
        "date": 0.60,
        "amount": 0.60,
        "currency": 0.60,
        "status": 0.72,
        "department": 0.72,
        "category": 0.72,
        "notes": 0.72,
    }

    assignments: dict[int, str] = {}
    confidences: dict[int, float] = {}
    taken_indices: set[int] = set()

    for role in ("name", "date", "amount", "currency", "status", "department", "category", "notes"):
        best_idx = None
        best_score = 0.0
        for idx, scores in candidate_scores.items():
            if idx in taken_indices:
                continue
            score = scores.get(role, 0.0)
            if score > best_score:
                best_idx = idx
                best_score = score
        if best_idx is not None and best_score >= thresholds[role]:
            assignments[best_idx] = role
            confidences[best_idx] = round(best_score, 2)
            taken_indices.add(best_idx)

    def assign_unique_detected_type(
        role: str, detected_type: str, minimum: float = 0.58
    ) -> None:
        if role in assignments.values():
            return
        candidates = [
            idx for idx, header in enumerate(headers)
            if idx not in taken_indices and columns.get(header, {}).get("detected_type") == detected_type
        ]
        if len(candidates) == 1:
            idx = candidates[0]
            score = max(minimum, candidate_scores[idx].get(role, 0.0))
            assignments[idx] = role
            confidences[idx] = round(score, 2)
            taken_indices.add(idx)

    assign_unique_detected_type("name", "name")
    assign_unique_detected_type("date", "date")
    assign_unique_detected_type("amount", "currency/amount")
    assign_unique_detected_type("currency", "currency code")
    assign_unique_detected_type("notes", "free text", minimum=0.55)

    primary_roles = set(assignments.values())
    enabled = (
        "amount" in primary_roles
        and len(primary_roles.intersection({"name", "date", "currency", "status", "department", "category"})) >= 2
    )
    if not enabled:
        return SemanticPlan(False, {}, {}, 0, None, None, None, [])

    label_idx = next(
        (idx for role_name in ("name", "department", "category", "notes") for idx, assigned in assignments.items() if assigned == role_name),
        0,
    )
    fill_down_indices = [
        idx for idx, role in assignments.items()
        if role in {"department", "category", "status", "currency"}
    ]

    return SemanticPlan(
        enabled=True,
        roles_by_index=assignments,
        confidence_by_index=confidences,
        label_idx=label_idx,
        amount_idx=next((idx for idx, role in assignments.items() if role == "amount"), None),
        currency_idx=next((idx for idx, role in assignments.items() if role == "currency"), None),
        date_idx=next((idx for idx, role in assignments.items() if role == "date"), None),
        fill_down_indices=fill_down_indices,
    )


def split_amount_currency_fields_dynamic(
    row: list[str], row_num: int, headers: list[str], amount_idx: int | None, currency_idx: int | None
) -> tuple[list[str], list[Change]]:
    if amount_idx is None or currency_idx is None:
        return row, []

    row = row.copy()
    changes: list[Change] = []
    amount_label = headers[amount_idx]
    currency_label = headers[currency_idx]

    amount_val = row[amount_idx]
    currency_val = row[currency_idx]

    extracted_amount, extracted_currency = extract_currency_from_text(amount_val)
    if extracted_currency and (not currency_val.strip()):
        if extracted_amount and extracted_amount != amount_val:
            row[amount_idx] = extracted_amount
            changes.append(
                Change(
                    row_num,
                    amount_label,
                    amount_val,
                    extracted_amount,
                    "Fixed",
                    "Currency marker removed from amount-like field so the numeric value can be parsed cleanly",
                )
            )
        row[currency_idx] = extracted_currency
        changes.append(
            Change(
                row_num,
                currency_label,
                currency_val,
                extracted_currency,
                "Fixed",
                "Currency recovered from amount-like field",
            )
        )
        amount_val = row[amount_idx]
        currency_val = row[currency_idx]

    if not row[amount_idx].strip() and currency_val.strip():
        extracted_amount, extracted_currency = extract_currency_from_text(currency_val)
        if extracted_amount:
            original_currency = row[currency_idx]
            row[amount_idx] = extracted_amount
            changes.append(
                Change(
                    row_num,
                    amount_label,
                    "",
                    extracted_amount,
                    "Fixed",
                    "Amount recovered from currency-like field",
                )
            )
            if extracted_currency:
                row[currency_idx] = extracted_currency
                changes.append(
                    Change(
                        row_num,
                        currency_label,
                        original_currency,
                        extracted_currency,
                        "Fixed",
                        "Currency standardised after recovering combined amount/currency text",
                    )
                )

    return row, changes


def apply_semantic_normalisations(
    row: list[str], row_num: int, headers: list[str], semantic_plan: SemanticPlan
) -> tuple[list[str], list[Change]]:
    row = row.copy()
    changes: list[Change] = []

    row, split_changes = split_amount_currency_fields_dynamic(
        row, row_num, headers, semantic_plan.amount_idx, semantic_plan.currency_idx
    )
    changes.extend(split_changes)

    for idx, role in semantic_plan.roles_by_index.items():
        original = row[idx]
        if role == "date":
            new_value, changed, reason = normalise_date(original)
        elif role == "amount":
            new_value, changed, reason = normalise_amount(original)
        elif role == "currency":
            new_value, changed, reason = normalise_currency(original)
        elif role == "name":
            new_value, changed, reason = normalise_name(original)
        elif role == "status":
            new_value, changed, reason = normalise_status(original)
        elif role in {"department", "category"}:
            new_value = " ".join(original.split()).title() if original.strip() else ""
            changed = new_value != original
            reason = f"{role.title()} title-cased"
        else:
            continue

        if changed:
            row[idx] = new_value
            changes.append(Change(row_num, headers[idx], original, new_value, "Fixed", reason))

    return row, changes


def needs_review_semantic(row: list[str], structure_changed: bool, semantic_plan: SemanticPlan) -> bool:
    if structure_changed:
        return True
    if semantic_plan.amount_idx is not None and not row[semantic_plan.amount_idx].strip():
        return True
    if semantic_plan.date_idx is not None:
        date_value = row[semantic_plan.date_idx].strip()
        if date_value and not re.match(r"^\d{4}-\d{2}-\d{2}$", date_value):
            return True
    review_tokens = {"nan", "null", "n/a", "na", "not applicable", "none", "inf"}
    return any(cell.strip().lower() in review_tokens for cell in row if cell.strip())


def needs_review_generic(row: list[str], structure_changed: bool) -> bool:
    if structure_changed:
        return True
    review_tokens = {"nan", "null", "n/a", "na", "not applicable", "none", "inf"}
    return any(cell.strip().lower() in review_tokens for cell in row if cell.strip())


def process_generic(
    all_rows: list[list[str]], delimiter: str, initial_changelog: list[Change] | None = None
) -> tuple[list[CleanRow], list[QuarantineRow], list[Change], list[str], str]:
    headers, header_changes = normalise_headers_generic(all_rows[0])
    n_cols = len(headers)
    header_sig = tuple(h.strip().lower() for h in headers)
    semantic_plan = build_semantic_plan(headers, all_rows[1:], delimiter)
    applied_mode = "semantic" if semantic_plan.enabled else "generic"

    clean_data: list[CleanRow] = []
    quarantine: list[QuarantineRow] = []
    changelog: list[Change] = list(initial_changelog or []) + list(header_changes)
    seen_exact: dict[tuple, int] = {}
    running_amount_total = 0.0
    amount_idx = semantic_plan.amount_idx
    if amount_idx is None:
        amount_idx = next((i for i, header in enumerate(headers) if "amount" in header.lower() or "total" in header.lower()), None)
    label_idx = semantic_plan.label_idx if semantic_plan.enabled else 0

    for i, raw_row in enumerate(all_rows[1:], start=2):
        cls = classify_raw_row_generic(raw_row, header_sig, n_cols)
        if cls != "NORMAL":
            q_reason = GENERIC_QUARANTINE_REASONS[cls]
            q_row = [c.strip() for c in raw_row]
            q_row = (q_row + [""] * n_cols)[:n_cols]
            row_id = next((c for c in q_row if c), "[empty]")
            column_hint = "[row]"
            if cls == "FORMULA":
                _, formula_column = detect_formula_row(raw_row, headers)
                column_hint = formula_column or "formula_residue"
            quarantine.append(QuarantineRow(q_row, i, q_reason))
            changelog.append(
                Change(
                    i,
                    column_hint,
                    row_id[:60],
                    "",
                    "Quarantined",
                    "formula_residue: Excel formula found, not data" if cls == "FORMULA" else q_reason,
                )
            )
            continue

        aligned, align_change, structure_changed = fix_alignment_generic(raw_row, i, n_cols, delimiter)
        if align_change:
            changelog.append(align_change)

        cleaned, cell_changes = clean_row_generic(aligned, i, headers)
        changelog.extend(cell_changes)
        semantic_changes: list[Change] = []
        if semantic_plan.enabled:
            cleaned, semantic_changes = apply_semantic_normalisations(cleaned, i, headers, semantic_plan)
            changelog.extend(semantic_changes)
        was_modified = bool(align_change or cell_changes or semantic_changes)

        label_text = cleaned[label_idx] if label_idx < len(cleaned) else ""
        amount_text = cleaned[amount_idx] if amount_idx is not None and amount_idx < len(cleaned) else ""
        if amount_idx is not None and (
            row_amount_totalish(label_text, amount_text, running_amount_total)
            or sparse_total_label_row(cleaned, label_idx, amount_idx)
        ):
            quarantine.append(QuarantineRow(cleaned, i, GENERIC_QUARANTINE_REASONS["CALCULATED_SUBTOTAL"]))
            changelog.append(Change(i, headers[amount_idx], amount_text, "", "Quarantined", "Calculated subtotal row"))
            continue

        row_key = tuple(cleaned)
        if row_key in seen_exact:
            first_row = seen_exact[row_key]
            changelog.append(
                Change(i, "[row]", cleaned[0] if cleaned else "", "", "Removed", f"Exact duplicate of row {first_row}")
            )
            continue
        seen_exact[row_key] = i

        clean_data.append(
            CleanRow(
                row=cleaned,
                row_num=i,
                was_modified=was_modified,
                needs_review=(
                    needs_review_semantic(cleaned, structure_changed, semantic_plan)
                    if semantic_plan.enabled
                    else needs_review_generic(cleaned, structure_changed)
                ),
            )
        )
        if amount_idx is not None:
            parsed_amount = parse_amount_like(cleaned[amount_idx])
            if parsed_amount is not None:
                running_amount_total += parsed_amount

    if semantic_plan.enabled and semantic_plan.fill_down_indices:
        forward_fill_merged_cell_gaps_generic(clean_data, changelog, headers, semantic_plan.fill_down_indices)
        flag_near_duplicates_semantic(clean_data, changelog, headers, semantic_plan)

    return clean_data, quarantine, changelog, headers, applied_mode


# ══════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════════

def _header_font() -> Font:
    return Font(bold=True, color="FFFFFF")

def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _style_sheet(ws, col_widths: list[int], header_color: str):
    """Apply bold header, color, frozen row, and column widths."""
    fill = _header_fill(header_color)
    font = _header_font()
    for cell in ws[1]:
        cell.font  = font
        cell.fill  = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.freeze_panes = "A2"
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _infer_col_widths(rows: list[list], min_width: int = 10, max_width: int = 60, sample: int = 300) -> list[int]:
    if not rows:
        return []
    widths = [max(min_width, min(max_width, len(str(v)) + 2)) for v in rows[0]]
    for row in rows[1 : sample + 1]:
        for i, val in enumerate(row):
            widths[i] = max(widths[i], min(max_width, len(str(val)) + 2))
    return [max(min_width, min(max_width, w)) for w in widths]

# Accent fills for was_modified / needs_review cells
FILL_MODIFIED = PatternFill("solid", fgColor="FFF2CC")   # soft yellow
FILL_REVIEW   = PatternFill("solid", fgColor="FCE4D6")   # soft orange

def write_workbook(
    clean_data:  list[CleanRow],
    quarantine:  list[QuarantineRow],
    changelog:   list[Change],
    output_path: Path,
    headers: list[str] | None = None,
) -> None:
    headers = headers or HEADERS
    wb = openpyxl.Workbook()

    # ── Sheet 1 — Clean Data ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Clean Data"
    clean_headers = headers + ["was_modified", "needs_review"]
    clean_rows_for_width = [clean_headers]
    ws1.append(clean_headers)
    for entry in clean_data:
        row_out = entry.row + [entry.was_modified, entry.needs_review]
        ws1.append(row_out)
        clean_rows_for_width.append(row_out)
        # Accent modified / review flag cells
        last = ws1.max_row
        mod_cell    = ws1.cell(last, len(headers) + 1)
        review_cell = ws1.cell(last, len(headers) + 2)
        if entry.was_modified:
            mod_cell.fill = FILL_MODIFIED
        if entry.needs_review:
            review_cell.fill = FILL_REVIEW
    _style_sheet(ws1, _infer_col_widths(clean_rows_for_width), "4CAF50")   # green

    # ── Sheet 2 — Quarantine ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Quarantine")
    quarantine_headers = headers + ["quarantine_reason"]
    quarantine_rows_for_width = [quarantine_headers]
    ws2.append(quarantine_headers)
    for q in quarantine:
        row_out = q.row + [q.reason]
        ws2.append(row_out)
        quarantine_rows_for_width.append(row_out)
    _style_sheet(ws2, _infer_col_widths(quarantine_rows_for_width), "E53935")   # red

    # ── Sheet 3 — Change Log ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Change Log")
    log_headers = ["original_row_number", "column_affected",
                   "original_value", "new_value", "action_taken", "reason"]
    log_rows_for_width = [log_headers]
    ws3.append(log_headers)
    for c in changelog:
        row_out = [c.original_row_number, c.column_affected,
                   c.original_value, c.new_value, c.action_taken, c.reason]
        ws3.append(row_out)
        log_rows_for_width.append(row_out)
    _style_sheet(ws3, _infer_col_widths(log_rows_for_width), "1565C0")   # blue

    # Notes column text-wrap in Sheet 1
    if "Notes" in headers:
        notes_col = get_column_letter(headers.index("Notes") + 1)
        for cell in ws1[notes_col][1:]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Reason column text-wrap in Sheet 3
    for cell in ws3["F"][1:]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ══════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════

ASSUMPTIONS = [
    "Rows that still contain Excel formulas as text (for example '=SUM(...)') are quarantined because they are not stable data values",
    "Ambiguous DD/MM vs MM/DD dates: day-first assumed; fallback to month-first if day-first is impossible",
    "MM-DD-YY two-digit years: treated as 2000–2049 for YY < 50, 1950–1999 for YY ≥ 50",
    "Unix timestamps: interpreted as UTC; converted to YYYY-MM-DD",
    "Excel serial dates: Windows epoch (1899-12-30); range 40,000–55,000 treated as dates",
    "European decimal 1.200,00: detected when period precedes comma; converted to 1200.00",
    "Combined values like '$1,200 USD' are split so Amount keeps the numeric value and Currency keeps the ISO code",
    "Blank / N/A / TBD amounts: cleared to empty string and flagged needs_review=TRUE",
    "\"INR ₹\" and similar symbol+code combos: symbol stripped, ISO code kept",
    "\"Eng\" department abbreviation: kept as-is (expanding abbreviations requires a lookup table)",
    "Short blank runs in categorical columns are forward-filled when they look like merged-cell export gaps; filled rows are flagged for review",
    "Rows with long single-cell prose are treated as notes/metadata rather than transactional data",
    "Rows labelled TOTAL/Subtotal/SUM are quarantined when the amount matches the running total closely enough to look calculated",
    "Near-duplicate rows (same Name/Amount/Currency/Category, date within 2 days): both kept, both flagged",
    "Exact duplicates: first occurrence kept; subsequent occurrences removed and logged",
    "Short rows (< 8 columns): padded with empty strings; flagged needs_review=TRUE",
    f"Metadata export row (sparse, 1/8 fields filled): quarantined as 'Less than {int(SPARSE_THRESHOLD_SCHEMA * 100)}% columns filled'",
]

GENERIC_ASSUMPTIONS = [
    "Rows that still contain Excel formulas as text are quarantined because they are not stable data values",
    "Delimiter is auto-detected from the file content (comma/semicolon/tab/pipe)",
    "Rows with overflow columns are repaired by merging overflow into the last column",
    "Rows with missing trailing columns are padded with empty strings",
    "Repeated header rows and subtotal/total structural rows are quarantined",
    "Long one-cell prose rows are treated as notes rather than tabular data",
    "Rows before a detected header are moved to File Metadata entries in the Change Log",
    "BOM/null bytes/line breaks/smart quotes are normalised in text cells",
    "Exact duplicate rows are removed (first occurrence kept)",
]

SEMANTIC_ASSUMPTIONS = GENERIC_ASSUMPTIONS + [
    "When headers are non-standard, likely semantic roles are inferred from the column values and header hints",
    "Date-like columns are normalised to YYYY-MM-DD when confidence is high enough",
    "Amount-like and currency-like columns are normalised even when their headers are not exact schema matches",
    "Status-like, department-like, and category-like columns are title-cased or canonicalised when their semantics are clear enough",
    "Near-duplicate detection uses inferred semantic key columns instead of fixed schema names",
]


def _normalise_header_for_match(value: str) -> str:
    return " ".join(value.strip().lower().split())


def is_schema_specific_header(header_row: list[str]) -> bool:
    if len(header_row) != N_COLS:
        return False
    cleaned = tuple(_normalise_header_for_match(c or "") for c in header_row)
    expected = tuple(_normalise_header_for_match(c) for c in HEADERS)
    return cleaned == expected


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Heal messy tabular files into a 3-sheet Excel workbook.")
    parser.add_argument("input", nargs="?", default=str(INPUT), help="Input file path")
    parser.add_argument("output", nargs="?", default=str(OUTPUT), help="Output .xlsx path")
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--sheet", dest="sheet_name", help="Workbook sheet name to heal")
    group.add_argument(
        "--all-sheets",
        dest="all_sheets",
        action="store_true",
        help="Consolidate all sheets before healing (only when columns are compatible)",
    )
    return parser.parse_args(argv)


def execute_healing(
    input_path: Path,
    *,
    sheet_name: str | None = None,
    consolidate_sheets: bool | None = None,
) -> dict:
    if not input_path.exists():
        raise FileNotFoundError(f"File not found: {input_path}")

    all_rows, delimiter = read_file(
        input_path,
        sheet_name=sheet_name,
        consolidate_sheets=consolidate_sheets,
    )
    if len(all_rows) < 2:
        raise ValueError("File is empty or has only a header.")

    original_total_in = len(all_rows)
    all_rows, metadata_changes = preprocess_rows(all_rows)
    if len(all_rows) < 2:
        raise ValueError("File is empty after metadata/header detection.")

    if is_schema_specific_header(all_rows[0]):
        mode = "schema-specific"
        clean_data, quarantine, changelog = process_schema_specific(
            all_rows,
            initial_changelog=metadata_changes,
        )
        headers = HEADERS
        assumptions = ASSUMPTIONS
    else:
        clean_data, quarantine, changelog, headers, mode = process_generic(
            all_rows,
            delimiter,
            initial_changelog=metadata_changes,
        )
        assumptions = SEMANTIC_ASSUMPTIONS if mode == "semantic" else GENERIC_ASSUMPTIONS

    action_counts = Counter(c.action_taken for c in changelog)
    quarantine_reason_counts = {
        reason: sum(1 for q in quarantine if q.reason == reason)
        for reason in {q.reason for q in quarantine}
    }

    return {
        "input_path": input_path,
        "delimiter": delimiter,
        "total_in": original_total_in,
        "mode": mode,
        "headers": headers,
        "assumptions": assumptions,
        "clean_data": clean_data,
        "quarantine": quarantine,
        "changelog": changelog,
        "action_counts": action_counts,
        "quarantine_reason_counts": dict(sorted(quarantine_reason_counts.items())),
    }


def main():
    args = parse_args(sys.argv[1:])
    input_path = Path(args.input)
    output_path = Path(args.output)
    try:
        result = execute_healing(
            input_path,
            sheet_name=args.sheet_name,
            consolidate_sheets=True if args.all_sheets else None,
        )
    except (FileNotFoundError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)

    write_workbook(
        result["clean_data"],
        result["quarantine"],
        result["changelog"],
        output_path,
        headers=result["headers"],
    )

    W = 60
    print()
    print("═" * W)
    print("  CSV Doctor  ·  Heal Report  (Excel output)")
    print("═" * W)
    print(f"  Input file   : {input_path.name}")
    print(f"  Output file  : {output_path.name}")
    print(f"  Mode         : {result['mode']}")
    print(f"  Delimiter    : {repr(result['delimiter'])}")
    print("─" * W)
    print(f"  Rows in      : {result['total_in']}  (incl. column header row)")
    print(f"  Clean Data   : {len(result['clean_data'])} rows")
    print(f"    · was_modified = TRUE  : {sum(1 for r in result['clean_data'] if r.was_modified)}")
    print(f"    · needs_review = TRUE  : {sum(1 for r in result['clean_data'] if r.needs_review)}")
    print(f"  Quarantine   : {len(result['quarantine'])} rows")
    for reason, rows in result["quarantine_reason_counts"].items():
        print(f"    · {reason:<40} {rows}")
    print(f"  Changes logged: {len(result['changelog'])}")
    print(f"    · Fixed       : {result['action_counts'].get('Fixed', 0)}")
    print(f"    · Quarantined : {result['action_counts'].get('Quarantined', 0)}")
    print(f"    · Removed     : {result['action_counts'].get('Removed', 0)}")
    print(f"    · Flagged     : {result['action_counts'].get('Flagged', 0)}")
    print("─" * W)
    print("  ASSUMPTIONS MADE:")
    for a in result["assumptions"]:
        print(f"    · {a}")
    print("═" * W)
    print()


if __name__ == "__main__":
    main()
