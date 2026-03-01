from __future__ import annotations

import csv
import io
import re
from collections import Counter
from pathlib import Path

import openpyxl
import pandas as pd

from loader import load_file
from heal_modules.shared import (
    COL,
    CURRENCY_SYMBOL_MAP,
    FORMULA_RE,
    HEADERS,
    N_COLS,
    NOTES_ROW_RE,
    SPARSE_THRESHOLD_GENERIC,
    SPARSE_THRESHOLD_SCHEMA,
    STATUS_VALUE_HINTS,
    TOTAL_LABEL_RE,
    Change,
    _strip_nulls,
    is_schema_specific_header,
)
from heal_modules.normalization import (
    _clean_cell_text,
    extract_currency_from_text,
    normalise_date,
    parse_amount_like,
)

def _non_empty_cells(row: list[str]) -> list[str]:
    cells = []
    for cell in row:
        cleaned = _strip_nulls(cell).strip()
        if cleaned:
            cells.append(cleaned)
    return cells


def _joined_row_text(row: list[str]) -> str:
    return " | ".join(_strip_nulls(cell).strip() for cell in row if _strip_nulls(cell).strip())


def _looks_like_header_row(row: list[str]) -> bool:
    if is_schema_specific_header(row):
        return True

    non_empty = [_strip_nulls(cell).strip() for cell in row if _strip_nulls(cell).strip()]
    if len(non_empty) < 2:
        return False
    if any(FORMULA_RE.match(cell) for cell in non_empty):
        return False
    if sum(1 for cell in non_empty if len(cell) > 50) >= 2:
        return False
    data_like_cells = 0
    for cell in non_empty:
        lower_cell = cell.strip().lower()
        if parse_amount_like(cell) is not None:
            data_like_cells += 1
            continue
        extracted_amount, _ = extract_currency_from_text(cell)
        if extracted_amount and parse_amount_like(extracted_amount) is not None:
            data_like_cells += 1
            continue
        normalized_date, changed, _ = normalise_date(cell)
        if changed or re.match(r"^\d{4}-\d{2}-\d{2}$", normalized_date):
            data_like_cells += 1
            continue
        if lower_cell in STATUS_VALUE_HINTS:
            data_like_cells += 1
    if data_like_cells >= 2:
        return False
    alpha_cells = sum(1 for cell in non_empty if re.search(r"[A-Za-z]", cell))
    numeric_heavy = sum(1 for cell in non_empty if re.fullmatch(r"[\d,./:-]+", cell))
    return alpha_cells >= max(2, len(non_empty) - 1) and numeric_heavy <= 1


DATE_SIGNAL_RE = re.compile(
    r"(\b\d{1,4}[/-]\d{1,2}[/-]\d{1,4}\b|\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\b)",
    re.IGNORECASE,
)


def _row_data_signal_count(row: list[str]) -> int:
    signals = 0
    for cell in _non_empty_cells(row):
        lowered = cell.lower()
        if parse_amount_like(cell) is not None:
            signals += 1
            continue
        extracted_amount, _ = extract_currency_from_text(cell)
        if extracted_amount and parse_amount_like(extracted_amount) is not None:
            signals += 1
            continue
        if lowered in STATUS_VALUE_HINTS:
            signals += 1
            continue
        if DATE_SIGNAL_RE.search(cell):
            signals += 1
    return signals


def detect_header_row_index(all_rows: list[list[str]], explicit_header_row: int | None = None) -> int:
    if explicit_header_row is not None:
        return max(0, min(len(all_rows) - 1, explicit_header_row - 1))

    search_rows = all_rows[:20]
    exact_matches = [
        idx for idx, row in enumerate(search_rows)
        if is_schema_specific_header(row) and idx < len(all_rows) - 1
    ]
    if exact_matches:
        return exact_matches[0]

    generic_candidates = [
        idx for idx, row in enumerate(search_rows)
        if idx < len(all_rows) - 1 and _looks_like_header_row(row)
    ]
    if generic_candidates:
        signal_candidates = [
            idx for idx in generic_candidates
            if _row_data_signal_count(all_rows[idx + 1]) > _row_data_signal_count(all_rows[idx])
            and _row_data_signal_count(all_rows[idx + 1]) > 0
        ]
        if signal_candidates:
            return signal_candidates[-1]
        return generic_candidates[-1]
    return 0


def detect_header_band_start_index(all_rows: list[list[str]], header_idx: int) -> int:
    band_start = header_idx
    max_band_rows = 4

    while band_start > 0 and (header_idx - band_start + 1) < max_band_rows:
        previous = all_rows[band_start - 1]
        non_empty = _non_empty_cells(previous)
        if len(non_empty) < 2:
            break
        if not _looks_like_header_row(previous):
            break
        band_start -= 1

    return band_start


def _expand_header_band_row(row: list[str], width: int) -> list[str]:
    padded = [(_strip_nulls(cell).strip()) for cell in row] + [""] * max(0, width - len(row))
    expanded: list[str] = []
    current = ""
    for cell in padded[:width]:
        if cell:
            current = cell
            expanded.append(cell)
        else:
            expanded.append(current)
    return expanded


def merge_header_band_rows(rows: list[list[str]]) -> list[str]:
    width = max((len(row) for row in rows), default=0)
    if width == 0:
        return []

    expanded_rows = [_expand_header_band_row(row, width) for row in rows]
    merged: list[str] = []
    for col_idx in range(width):
        tokens: list[str] = []
        seen: set[str] = set()
        for row in expanded_rows:
            value = row[col_idx].strip()
            if not value:
                continue
            key = value.lower()
            if key in seen:
                continue
            tokens.append(value)
            seen.add(key)
        merged.append(" ".join(tokens).strip())
    return merged


def trim_sparse_edge_columns(all_rows: list[list[str]]) -> tuple[list[list[str]], list[Change]]:
    if not all_rows:
        return all_rows, []

    width = max((len(row) for row in all_rows), default=0)
    if width == 0:
        return all_rows, []

    padded = [row + [""] * (width - len(row)) for row in all_rows]
    data_rows = padded[1:] if len(padded) > 1 else []
    data_threshold = max(1, int(len(data_rows) * 0.15)) if data_rows else 1

    def non_empty_count(col_idx: int) -> int:
        return sum(1 for row in padded if row[col_idx].strip())

    left = 0
    while left < width:
        header_cell = padded[0][left].strip()
        if header_cell:
            break
        if non_empty_count(left) > data_threshold:
            break
        left += 1

    right = width
    while right > left:
        header_cell = padded[0][right - 1].strip()
        if header_cell:
            break
        if non_empty_count(right - 1) > data_threshold:
            break
        right -= 1

    if left == 0 and right == width:
        return all_rows, []

    trimmed = [row[left:right] for row in padded]
    changes: list[Change] = []
    if left > 0:
        changes.append(
            Change(
                1,
                "[column trimming]",
                f"Removed {left} leading sparse column(s)",
                "",
                "Fixed",
                "Sparse leading workbook columns removed before semantic/header detection",
            )
        )
    if right < width:
        changes.append(
            Change(
                1,
                "[column trimming]",
                f"Removed {width - right} trailing sparse column(s)",
                "",
                "Fixed",
                "Sparse trailing workbook columns removed before semantic/header detection",
            )
        )
    return trimmed, changes


def preprocess_rows(
    all_rows: list[list[str]],
    *,
    explicit_header_row: int | None = None,
) -> tuple[list[list[str]], list[Change]]:
    header_idx = detect_header_row_index(all_rows, explicit_header_row=explicit_header_row)
    if header_idx <= 0:
        trimmed_rows, trim_changes = trim_sparse_edge_columns(all_rows)
        return trimmed_rows, trim_changes

    header_band_start = detect_header_band_start_index(all_rows, header_idx)
    changes: list[Change] = []
    for i, row in enumerate(all_rows[:header_band_start], start=1):
        row_text = _joined_row_text(row) or "[empty metadata row]"
        changes.append(
            Change(
                i,
                "[file metadata]",
                row_text[:200],
                "",
                "Removed",
                "File Metadata: row before detected header moved out of the dataset",
            )
        )

    header_rows = all_rows[header_band_start : header_idx + 1]
    if len(header_rows) > 1:
        merged_header = merge_header_band_rows(header_rows)
        changes.append(
            Change(
                header_band_start + 1,
                "[header band]",
                " | ".join(_joined_row_text(row) for row in header_rows if _joined_row_text(row))[:200],
                " | ".join(cell for cell in merged_header if cell)[:200],
                "Fixed",
                "Multi-row workbook header band merged into a single semantic header row",
            )
        )
        trimmed_rows, trim_changes = trim_sparse_edge_columns([merged_header] + all_rows[header_idx + 1 :])
        return trimmed_rows, changes + trim_changes

    trimmed_rows, trim_changes = trim_sparse_edge_columns(all_rows[header_idx:])
    return trimmed_rows, changes + trim_changes


def is_formula_residue(value: str) -> bool:
    return bool(FORMULA_RE.match(_strip_nulls(value).strip()))


def detect_formula_row(row: list[str], headers: list[str]) -> tuple[bool, str]:
    for idx, cell in enumerate(row):
        if is_formula_residue(cell):
            label = headers[idx] if idx < len(headers) else f"[col {idx + 1}]"
            return True, label
    return False, ""


def looks_like_notes_row(row: list[str]) -> bool:
    non_empty = _non_empty_cells(row)
    if len(non_empty) != 1:
        return False
    text = non_empty[0]
    if len(text) <= 50:
        return False
    if len(text.split()) < 8:
        return False
    return bool(NOTES_ROW_RE.search(text) or re.search(r"[A-Za-z]{4,}", text))


def row_amount_totalish(label_cell: str, amount_cell: str, running_total: float) -> bool:
    if not TOTAL_LABEL_RE.search(label_cell or ""):
        return False
    amount = parse_amount_like(amount_cell or "")
    if amount is None:
        return False
    tolerance = max(1.0, abs(running_total) * 0.02)
    return abs(amount - running_total) <= tolerance


def sparse_total_label_row(row: list[str], label_idx: int, amount_idx: int) -> bool:
    label_cell = row[label_idx] if label_idx < len(row) else ""
    amount_cell = row[amount_idx] if amount_idx < len(row) else ""
    if not TOTAL_LABEL_RE.search(label_cell or ""):
        return False
    if parse_amount_like(amount_cell or "") is None:
        return False
    non_empty = sum(1 for cell in row if cell.strip())
    return non_empty <= 2


def read_file(
    path: Path,
    *,
    sheet_name: str | None = None,
    consolidate_sheets: bool | None = None,
) -> tuple[list[list[str]], str]:
    """
    Load any supported file format via loader.load_file().

    Returns (raw_rows, delimiter).  For non-text formats (Excel, ODS, JSON)
    raw_rows are reconstructed from the DataFrame so the rest of the
    processing pipeline stays unchanged.
    """
    result    = load_file(path, sheet_name=sheet_name, consolidate_sheets=consolidate_sheets)
    raw_text  = result["raw_text"]
    delimiter = result["delimiter"]

    if raw_text is not None and delimiter is not None:
        # Text format: re-parse with csv.reader so multi-line quoted fields
        # are reconstructed correctly (same behaviour as the old read_mixed_encoding).
        rows = list(csv.reader(io.StringIO(raw_text), delimiter=delimiter))
    else:
        # Structured/binary formats: preserve actual workbook rows for healing.
        # Reconstructing from DataFrame headers loses workbook preambles and
        # misclassifies real header rows when metadata bands precede the table.
        suffix = path.suffix.lower()
        delimiter = ","
        if suffix in {".xlsx", ".xlsm"}:
            workbook = openpyxl.load_workbook(path, data_only=False)
            if consolidate_sheets:
                rows = _read_openpyxl_rows_consolidated(workbook, result.get("sheet_names") or workbook.sheetnames)
            else:
                active_sheet = result.get("sheet_name")
                if not active_sheet or str(active_sheet).startswith("[all "):
                    active_sheet = workbook.sheetnames[0]
                rows = _read_openpyxl_rows(workbook[active_sheet])
        elif suffix in {".xls", ".ods"}:
            rows = _read_spreadsheet_rows_with_pandas(
                path,
                suffix=suffix,
                sheet_name=None if consolidate_sheets else result.get("sheet_name"),
                all_sheets=result.get("sheet_names"),
                consolidate_sheets=bool(consolidate_sheets),
            )
        else:
            # JSON and other structured formats do not have workbook-style preambles,
            # so DataFrame reconstruction remains acceptable here.
            df = result["dataframe"]
            rows = [list(df.columns)] + [
                [str(v) if v is not None else "" for v in row]
                for row in df.itertuples(index=False, name=None)
            ]

    return rows, delimiter or ","


def _trim_trailing_empty_cells(row: list[str]) -> list[str]:
    trimmed = list(row)
    while trimmed and not str(trimmed[-1]).strip():
        trimmed.pop()
    return trimmed


def _normalise_row_cells(values) -> list[str]:
    row = ["" if value is None else str(value) for value in values]
    return _trim_trailing_empty_cells(row)


def _read_openpyxl_rows(sheet) -> list[list[str]]:
    rows: list[list[str]] = []
    for values in sheet.iter_rows(values_only=True):
        rows.append(_normalise_row_cells(values))
    return rows


def _read_openpyxl_rows_consolidated(workbook, sheet_names: list[str]) -> list[list[str]]:
    all_rows: list[list[str]] = []
    header_added = False
    for name in sheet_names:
        sheet_rows = _read_openpyxl_rows(workbook[name])
        if not sheet_rows:
            continue
        if not header_added:
            all_rows.extend(sheet_rows)
            header_added = True
        else:
            all_rows.extend(sheet_rows[1:])
    return all_rows


def _read_spreadsheet_rows_with_pandas(
    path: Path,
    *,
    suffix: str,
    sheet_name: str | None,
    all_sheets: list[str] | None,
    consolidate_sheets: bool,
) -> list[list[str]]:
    engine = "odf" if suffix == ".ods" else None
    read_kwargs = {"header": None, "dtype": str}
    if engine:
        read_kwargs["engine"] = engine

    def frame_to_rows(df: pd.DataFrame) -> list[list[str]]:
        return [
            _trim_trailing_empty_cells(["" if value is None else str(value) for value in row])
            for row in df.fillna("").itertuples(index=False, name=None)
        ]

    if consolidate_sheets:
        rows: list[list[str]] = []
        names = all_sheets or []
        header_added = False
        for name in names:
            df = pd.read_excel(path, sheet_name=name, **read_kwargs)
            sheet_rows = frame_to_rows(df)
            if not sheet_rows:
                continue
            if not header_added:
                rows.extend(sheet_rows)
                header_added = True
            else:
                rows.extend(sheet_rows[1:])
        return rows

    df = pd.read_excel(path, sheet_name=sheet_name, **read_kwargs)
    return frame_to_rows(df)


# ══════════════════════════════════════════════════════════════════════════
# STEP 2 — Row classification
# ══════════════════════════════════════════════════════════════════════════

QUARANTINE_REASONS = {
    "EMPTY":               "Completely empty row",
    "WHITESPACE":          "Row is all whitespace",
    "STRUCTURAL_HEADER":   "Structural row (TOTAL/subtotal/header repeat)",
    "STRUCTURAL_TOTAL":    "Structural row (TOTAL/subtotal/header repeat)",
    "CALCULATED_SUBTOTAL": "Calculated subtotal row",
    "NOTES_ROW":           "Appears to be a notes row",
    "FORMULA":             "Excel formula found, not data",
    "SPARSE":              f"Less than {int(SPARSE_THRESHOLD_SCHEMA * 100)}% columns filled",
}

def classify_raw_row(row: list[str], header_sig: tuple[str, ...]) -> str:
    stripped = [c.strip() for c in row]

    if all(c == "" for c in stripped):
        # Distinguish truly empty (all "") from whitespace-only
        return "WHITESPACE" if any(c != "" for c in row) else "EMPTY"

    if tuple(c.lower() for c in stripped) == header_sig:
        return "STRUCTURAL_HEADER"

    if looks_like_notes_row(row):
        return "NOTES_ROW"

    if any(is_formula_residue(cell) for cell in stripped):
        return "FORMULA"

    if TOTAL_LABEL_RE.search(stripped[0]) and len(stripped) > COL["Amount"]:
        if parse_amount_like(stripped[COL["Amount"]]) is not None:
            return "NORMAL"

    non_empty = sum(1 for c in stripped if c)
    if non_empty < N_COLS * SPARSE_THRESHOLD_SCHEMA:
        return "SPARSE"

    return "NORMAL"


# ══════════════════════════════════════════════════════════════════════════
# STEP 3 — Alignment fix (with Change log entry)
# ══════════════════════════════════════════════════════════════════════════

def fix_alignment(row: list[str], row_num: int) -> tuple[list[str], Change | None]:
    """Detect and fix three structural column problems."""
    n = len(row)
    if n == N_COLS:
        return row, None

    if n > N_COLS:
        # Shifted right: leading empty ghost column
        if row[0].strip() == "" and row[1].strip() != "":
            fixed = row[1: N_COLS + 1]
            return fixed, Change(
                row_num, "[row structure]",
                f"{n} columns (empty leading ghost col)",
                f"{N_COLS} columns",
                "Fixed", "Shifted-right row: empty leading column stripped"
            )
        # Phantom comma: empty ghost field sits between Status and Notes
        if n == N_COLS + 1 and row[N_COLS - 1].strip() == "" and row[N_COLS].strip() != "":
            fixed = row[: N_COLS - 1] + [row[N_COLS]]
            return fixed, Change(
                row_num, "Notes",
                f"[ghost field] + '{row[N_COLS]}'",
                row[N_COLS],
                "Fixed", "Phantom comma: empty ghost field before Notes removed"
            )
        # Unquoted commas in Notes: merge overflow columns back
        merged = ", ".join(row[N_COLS - 1:]).strip()
        fixed = row[: N_COLS - 1] + [merged]
        return fixed, Change(
            row_num, "Notes",
            f"{n - N_COLS + 1} fragments: {row[N_COLS - 1]!r}...",
            merged,
            "Fixed", "Unquoted commas in Notes field: fragments merged into one"
        )

    # Short row: pad with empty strings
    fixed = row + [""] * (N_COLS - n)
    return fixed, Change(
        row_num, "[row structure]",
        f"{n} columns",
        f"{N_COLS} columns ({N_COLS - n} empty field(s) appended)",
        "Fixed", f"Short row padded with {N_COLS - n} empty field(s)"
    )

GENERIC_QUARANTINE_REASONS = {
    "EMPTY":             "Completely empty row",
    "WHITESPACE":        "Row is all whitespace",
    "STRUCTURAL_HEADER": "Structural row (header repeated)",
    "STRUCTURAL_TOTAL":  "Structural row (TOTAL/subtotal)",
    "CALCULATED_SUBTOTAL": "Calculated subtotal row",
    "NOTES_ROW":         "Appears to be a notes row",
    "FORMULA":           "Excel formula found, not data",
    "SPARSE":            f"Less than {int(SPARSE_THRESHOLD_GENERIC * 100)}% columns filled",
}


def _normalise_header_text(value: str, index: int) -> str:
    cleaned = " ".join(value.replace("\ufeff", "").strip().split())
    if not cleaned:
        return f"column_{index}"
    return cleaned


def normalise_headers_generic(raw_header: list[str]) -> tuple[list[str], list[Change]]:
    headers = []
    changes: list[Change] = []
    seen = Counter()

    for i, cell in enumerate(raw_header, start=1):
        original = cell or ""
        cleaned, clean_reasons = _clean_cell_text(original)
        base = _normalise_header_text(cleaned, i)
        key = base.lower()
        seen[key] += 1
        final = f"{base}_{seen[key]}" if seen[key] > 1 else base

        reasons = list(clean_reasons)
        if seen[key] > 1:
            reasons.append("Duplicate header renamed with suffix")
        if final != original:
            reason = "; ".join(reasons) if reasons else "Header normalised"
            changes.append(Change(1, f"[header col {i}]", original, final, "Fixed", reason))

        headers.append(final)

    return headers, changes


def classify_raw_row_generic(row: list[str], header_sig: tuple[str, ...], n_cols: int) -> str:
    stripped = [c.strip() for c in row]

    if all(c == "" for c in stripped):
        return "WHITESPACE" if any(c != "" for c in row) else "EMPTY"

    if tuple(c.lower() for c in stripped) == header_sig:
        return "STRUCTURAL_HEADER"

    if looks_like_notes_row(row):
        return "NOTES_ROW"

    if any(is_formula_residue(cell) for cell in stripped):
        return "FORMULA"

    first_non_empty = next((c for c in stripped if c), "")
    non_empty = sum(1 for c in stripped if c)

    if re.match(r"^(grand\s+total|subtotal|total)\b", first_non_empty, flags=re.IGNORECASE):
        if non_empty <= max(2, n_cols // 3):
            return "STRUCTURAL_TOTAL"

    if non_empty < max(1, int(n_cols * SPARSE_THRESHOLD_GENERIC)):
        return "SPARSE"

    return "NORMAL"


def fix_alignment_generic(
    row: list[str], row_num: int, n_cols: int, delimiter: str
) -> tuple[list[str], Change | None, bool]:
    n = len(row)
    if n == n_cols:
        return row, None, False

    if n > n_cols:
        merged_tail = f"{delimiter} ".join(part for part in row[n_cols - 1:] if part != "")
        fixed = row[: n_cols - 1] + [merged_tail]
        return fixed, Change(
            row_num,
            "[row structure]",
            f"{n} columns",
            f"{n_cols} columns",
            "Fixed",
            f"Overflow columns merged into last column using delimiter '{delimiter}'",
        ), True

    fixed = row + [""] * (n_cols - n)
    return fixed, Change(
        row_num,
        "[row structure]",
        f"{n} columns",
        f"{n_cols} columns ({n_cols - n} empty field(s) appended)",
        "Fixed",
        f"Short row padded with {n_cols - n} empty field(s)",
    ), True


def clean_row_generic(
    row: list[str], row_num: int, headers: list[str]
) -> tuple[list[str], list[Change]]:
    cleaned = []
    changes: list[Change] = []

    for i, cell in enumerate(row):
        original = cell
        new_val, reasons = _clean_cell_text(cell)
        if reasons:
            col_label = headers[i] if i < len(headers) else f"[col {i + 1}]"
            orig_display = original.replace("\ufeff", "[BOM]").replace("\x00", "[NULL]")
            changes.append(
                Change(row_num, col_label, orig_display, new_val, "Fixed", "; ".join(reasons))
            )
        cleaned.append(new_val)

    return cleaned, changes
