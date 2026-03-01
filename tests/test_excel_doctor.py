from __future__ import annotations

import importlib.util
import io
import json
import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
EXCEL_SCRIPTS = ROOT / "skills" / "excel-doctor" / "scripts"
WEB_APP = ROOT / "web" / "app.py"
FIXTURE_DIR = ROOT / "tests" / "fixtures" / "excel"
LOADER_FIXTURE_DIR = ROOT / "tests" / "fixtures" / "loader"


def load_module(module_path: Path, module_name: str):
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


EXCEL_DIAGNOSE = load_module(EXCEL_SCRIPTS / "diagnose.py", "excel_doctor_diagnose_tests")
EXCEL_HEAL = load_module(EXCEL_SCRIPTS / "heal.py", "excel_doctor_heal_tests")
WEB_APP_MODULE = load_module(WEB_APP, "sheet_doctor_web_app_tests")


class ExcelDoctorTests(unittest.TestCase):
    def test_hidden_sheet_fixture_reports_hidden_and_very_hidden(self):
        report = EXCEL_DIAGNOSE.build_report(FIXTURE_DIR / "hidden_layers.xlsx")
        self.assertEqual(report["workbook_mode"], "workbook-native")
        self.assertEqual([item["name"] for item in report["sheets"]["hidden"]], ["Hidden"])
        self.assertEqual([item["name"] for item in report["sheets"]["very_hidden"]], ["VeryHidden"])

    def test_xlsm_fixture_is_supported_for_workbook_native_diagnose_and_heal(self):
        fixture = LOADER_FIXTURE_DIR / "multisheet.xlsm"
        report = EXCEL_DIAGNOSE.build_report(fixture)
        self.assertEqual(report["file_type"], ".xlsm")
        self.assertGreaterEqual(report["sheets"]["count"], 2)
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "multisheet_healed.xlsm"
            _, stats = EXCEL_HEAL.execute_healing(fixture, output_path)
            self.assertTrue(output_path.exists())
            self.assertEqual(output_path.suffix, ".xlsm")
            self.assertGreaterEqual(stats["sheets_processed"], 1)

    def test_stacked_header_fixture_reports_header_band_and_metadata(self):
        report = EXCEL_DIAGNOSE.build_report(FIXTURE_DIR / "stacked_headers.xlsx")
        self.assertIn("Report", report["header_bands"])
        self.assertEqual(report["header_bands"]["Report"]["rows"], [3, 4])
        self.assertEqual(report["metadata_rows"]["Report"], [1, 2])

    def test_preamble_fixture_reports_metadata_rows(self):
        report = EXCEL_DIAGNOSE.build_report(FIXTURE_DIR / "preamble_report.xlsx")
        self.assertEqual(report["metadata_rows"]["Export"], [1, 2])
        self.assertEqual(report["sheet_summaries"]["Export"]["header_row"], 3)

    def test_formula_fixture_reports_formula_cells_cache_misses_and_structural_rows(self):
        report = EXCEL_DIAGNOSE.build_report(FIXTURE_DIR / "formula_cases.xlsx")
        calc_formulas = report["formula_cells"]["Calc"]
        self.assertTrue(any(item["cell"] == "D2" for item in calc_formulas))
        self.assertTrue(any(item["cell"] == "D3" for item in calc_formulas))
        self.assertIn("Calc", report["formula_cache_misses"])
        self.assertIn("Calc", report["structural_rows"])
        self.assertTrue(any(row["label"] == "TOTAL" for row in report["structural_rows"]["Calc"]))
        self.assertTrue(any("does not repair broken formulas" in warning for warning in report["manual_review_warnings"]))

    def test_duplicate_header_fixture_reports_duplicates_and_mixed_types(self):
        report = EXCEL_DIAGNOSE.build_report(FIXTURE_DIR / "duplicate_headers.xlsx")
        self.assertEqual(report["duplicate_headers"]["Dupes"], ["customer_id"])
        self.assertIn("amount", report["mixed_types"]["Dupes"])

    def test_notes_totals_fixture_reports_notes_and_totals(self):
        report = EXCEL_DIAGNOSE.build_report(FIXTURE_DIR / "notes_totals.xlsx")
        self.assertIn("Ledger", report["notes_rows"])
        self.assertIn("Ledger", report["structural_rows"])
        self.assertTrue(any(row["label"] == "TOTAL" for row in report["structural_rows"]["Ledger"]))

    def test_ragged_clinical_fixture_reports_metadata_header_band_and_edge_columns(self):
        report = EXCEL_DIAGNOSE.build_report(FIXTURE_DIR / "ragged_clinical.xlsx")
        self.assertEqual(report["metadata_rows"]["Clinical"], [1, 2])
        self.assertEqual(report["sheet_summaries"]["Clinical"]["header_row"], 3)
        self.assertIn("Clinical", report["empty_edge_columns"])
        self.assertTrue(any("Header-band and metadata-row detection is heuristic" in warning for warning in report["manual_review_warnings"]))

    def test_heal_unmerges_ranges_flattens_headers_and_adds_change_log(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "stacked_headers_healed.xlsx"
            changes, stats = EXCEL_HEAL.execute_healing(FIXTURE_DIR / "stacked_headers.xlsx", output_path)
            workbook = load_workbook(output_path)
            sheet = workbook["Report"]
            self.assertEqual(sheet["A1"].value, "Employee ID")
            self.assertEqual(sheet["B1"].value, "Employee Name")
            self.assertIn("Change Log", workbook.sheetnames)
            self.assertGreater(stats["metadata_rows_removed"], 0)
            self.assertGreater(stats["header_bands_flattened"], 0)
            self.assertGreater(len(changes), 0)

    def test_heal_trims_edge_columns_unmerges_cells_and_normalises_dates(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "merged_edges_healed.xlsx"
            _, stats = EXCEL_HEAL.execute_healing(FIXTURE_DIR / "merged_edges.xlsx", output_path)
            workbook = load_workbook(output_path)
            sheet = workbook["Orders"]
            self.assertEqual(sheet.max_column, 3)
            self.assertFalse(sheet.merged_cells.ranges)
            self.assertEqual(sheet["B2"].value, "Acme Corp")
            self.assertEqual(sheet["B3"].value, "Acme Corp")
            self.assertGreater(stats["merged_ranges_unmerged"], 0)
            self.assertGreater(stats["edge_columns_trimmed"], 0)
            self.assertGreater(stats["empty_rows_removed"], 0)

    def test_heal_normalises_text_dates_and_removes_empty_rows(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "text_date_cleanup_healed.xlsx"
            _, stats = EXCEL_HEAL.execute_healing(FIXTURE_DIR / "text_date_cleanup.xlsx", output_path)
            workbook = load_workbook(output_path)
            sheet = workbook["TextDates"]
            self.assertEqual(sheet["A2"].value, 'Smart "quote" text')
            self.assertEqual(sheet["B2"].value, "2024-04-03")
            self.assertEqual(sheet["B3"].value, "2024-05-06")
            self.assertGreater(stats["dates_normalised"], 0)

    def test_structured_summary_reports_workbook_native_mode(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "hidden_layers_healed.xlsx"
            changes, stats = EXCEL_HEAL.execute_healing(FIXTURE_DIR / "hidden_layers.xlsx", output_path)
            summary = EXCEL_HEAL.build_structured_summary(
                input_path=FIXTURE_DIR / "hidden_layers.xlsx",
                output_path=output_path,
                changes=changes,
                stats=stats,
            )
        self.assertEqual(summary["mode"], "workbook-native")
        self.assertEqual(summary["run_summary"]["metrics"]["mode"], "workbook-native")

    def test_structured_summary_includes_formula_preservation_warning(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "formula_cases_healed.xlsx"
            changes, stats = EXCEL_HEAL.execute_healing(FIXTURE_DIR / "formula_cases.xlsx", output_path)
            summary = EXCEL_HEAL.build_structured_summary(
                input_path=FIXTURE_DIR / "formula_cases.xlsx",
                output_path=output_path,
                changes=changes,
                stats=stats,
            )
        self.assertGreater(stats["formula_cells_preserved"], 0)
        self.assertTrue(any("does not recalculate formulas" in warning for warning in summary["warnings"]))

    def test_execute_healing_is_atomic_on_failure(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = FIXTURE_DIR / "hidden_layers.xlsx"
            output_path = Path(tmpdir) / "atomic.xlsx"
            output_path.write_bytes(b"sentinel")
            workbook = load_workbook(input_path)

            def boom(_target):
                raise RuntimeError("save exploded")

            workbook.save = boom
            with mock.patch.object(EXCEL_HEAL, "load_workbook", return_value=workbook):
                with self.assertRaises(RuntimeError):
                    EXCEL_HEAL.execute_healing(input_path, output_path)
            self.assertEqual(output_path.read_bytes(), b"sentinel")

    def test_corrupt_workbook_raises_clear_value_error(self):
        with self.assertRaisesRegex(ValueError, "Could not read workbook"):
            EXCEL_DIAGNOSE.build_report(LOADER_FIXTURE_DIR / "corrupt.xlsx")

    def test_encrypted_workbook_raises_clear_value_error(self):
        with self.assertRaisesRegex(ValueError, "Password-protected / encrypted OOXML workbooks are not supported"):
            EXCEL_DIAGNOSE.build_report(LOADER_FIXTURE_DIR / "encrypted.xlsx")
        with self.assertRaisesRegex(ValueError, "Password-protected / encrypted OOXML workbooks are not supported"):
            EXCEL_HEAL.execute_healing(LOADER_FIXTURE_DIR / "encrypted.xlsx", Path(tempfile.gettempdir()) / "unused.xlsx")

    def test_cli_rejects_xls_with_explicit_message(self):
        buffer = io.StringIO()
        with mock.patch.object(sys, "argv", ["diagnose.py", str(LOADER_FIXTURE_DIR / "corrupt.xls")]):
            with redirect_stdout(buffer):
                with self.assertRaises(SystemExit) as exc:
                    EXCEL_DIAGNOSE.main()
        self.assertEqual(exc.exception.code, 1)
        self.assertIn(".xls is not supported by excel-doctor", buffer.getvalue())

    def test_ui_mode_details_distinguish_workbook_native_from_tabular_rescue(self):
        native = WEB_APP_MODULE.workbook_mode_details(".xlsx", tabular_rescue=False)
        tabular = WEB_APP_MODULE.workbook_mode_details(".xlsx", tabular_rescue=True)
        legacy = WEB_APP_MODULE.workbook_mode_details(".xls", tabular_rescue=True)
        self.assertEqual(native["mode"], "workbook-native")
        self.assertEqual(tabular["mode"], "tabular-rescue")
        self.assertEqual(legacy["mode"], "tabular-rescue-fallback")
        self.assertIn("flattened", tabular["tradeoff"])
        self.assertIn("preserving workbook", native["why"])

    def test_xlsm_summary_warns_about_macro_preservation_limit(self):
        fixture = LOADER_FIXTURE_DIR / "multisheet.xlsm"
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "multisheet_healed.xlsm"
            changes, stats = EXCEL_HEAL.execute_healing(fixture, output_path)
            summary = EXCEL_HEAL.build_structured_summary(
                input_path=fixture,
                output_path=output_path,
                changes=changes,
                stats=stats,
            )
        self.assertTrue(any("macro preservation" in warning for warning in summary["warnings"]))


if __name__ == "__main__":
    unittest.main()
