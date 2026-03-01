import importlib.util
import io
import tempfile
import unittest
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path
from typing import Optional
from unittest import mock

from openpyxl import Workbook

_XLRD_AVAILABLE = importlib.util.find_spec("xlrd") is not None
_ODFPY_AVAILABLE = importlib.util.find_spec("odf") is not None


REPO_ROOT = Path(__file__).resolve().parents[1]
LOADER_PATH = REPO_ROOT / "skills" / "csv-doctor" / "scripts" / "loader.py"


def load_loader_module():
    spec = importlib.util.spec_from_file_location("sheet_doctor_loader", LOADER_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def public_fixture(*candidates: str) -> Optional[Path]:
    for candidate in candidates:
        path = Path(candidate)
        if path.exists():
            return path
    return None


class LoaderLocalBehaviorTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.loader = load_loader_module()

    def test_plain_txt_is_rejected(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "notes.txt"
            path.write_text("This is a text file.\n", encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "does not appear to contain delimited/tabular data"):
                self.loader.load_file(path)

    def test_missing_xlrd_raises_clear_importerror(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "legacy.xls"
            path.write_bytes(b"not-a-real-xls")

            original_import = __import__

            def fake_import(name, globals=None, locals=None, fromlist=(), level=0):
                if name == "xlrd":
                    raise ImportError("simulated missing xlrd")
                return original_import(name, globals, locals, fromlist, level)

            with mock.patch("builtins.__import__", side_effect=fake_import):
                with self.assertRaisesRegex(ImportError, r"\.xls files require xlrd"):
                    self.loader.load_file(path)

    def test_missing_odfpy_raises_clear_importerror(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "sheet.ods"
            path.write_bytes(b"not-a-real-ods")

            original_import = __import__

            def fake_import(name, globals=None, locals=None, fromlist=(), level=0):
                if name == "odf":
                    raise ImportError("simulated missing odf")
                return original_import(name, globals, locals, fromlist, level)

            with mock.patch("builtins.__import__", side_effect=fake_import):
                with self.assertRaisesRegex(ImportError, r"\.ods files require odfpy"):
                    self.loader.load_file(path)

    def test_large_text_inputs_emit_degraded_mode_warning(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "large.csv"
            path.write_text("name,value\nAda,10\nGrace,11\n", encoding="utf-8")

            with mock.patch.object(self.loader, "LARGE_FILE_WARNING_BYTES", 1), \
                 mock.patch.object(self.loader, "LARGE_FILE_DEGRADED_BYTES", 1), \
                 mock.patch.object(self.loader, "LARGE_FILE_HARD_LIMIT_BYTES", 10_000), \
                 mock.patch.object(self.loader, "LARGE_ROW_WARNING_COUNT", 1_000), \
                 mock.patch.object(self.loader, "LARGE_ROW_DEGRADED_COUNT", 2_000), \
                 mock.patch.object(self.loader, "LARGE_ROW_HARD_LIMIT_COUNT", 10_000):
                result = self.loader.load_file(path)

            self.assertTrue(result["degraded_mode"]["active"])
            self.assertTrue(any("Degraded mode active" in warning for warning in result["warnings"]))
            self.assertTrue(any("Large file size" in reason for reason in result["degraded_mode"]["reasons"]))

    def test_hard_limit_rejects_oversized_input(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "tiny.csv"
            path.write_text("a,b\n1,2\n", encoding="utf-8")

            with mock.patch.object(self.loader, "LARGE_FILE_HARD_LIMIT_BYTES", 1):
                with self.assertRaisesRegex(ValueError, "too large for safe in-memory processing"):
                    self.loader.load_file(path)

    def test_multisheet_xlsx_requires_explicit_selection_noninteractive(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "multi.xlsx"
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Visible"
            ws1.append(["name", "score"])
            ws1.append(["Ada", 10])
            ws2 = wb.create_sheet("Backup")
            ws2.append(["name", "score"])
            ws2.append(["Grace", 11])
            wb.save(path)

            fake_stdin = mock.Mock()
            fake_stdin.isatty.return_value = False

            with mock.patch.object(self.loader.sys, "stdin", fake_stdin):
                with self.assertRaisesRegex(ValueError, "Multiple sheets found"):
                    self.loader.load_file(path)

    def test_multisheet_xlsx_loads_when_sheet_name_is_explicit(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "multi.xlsx"
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Visible"
            ws1.append(["name", "score"])
            ws1.append(["Ada", 10])
            ws2 = wb.create_sheet("Backup")
            ws2.append(["name", "score"])
            ws2.append(["Grace", 11])
            wb.save(path)

            fake_stdin = mock.Mock()
            fake_stdin.isatty.return_value = False

            with mock.patch.object(self.loader.sys, "stdin", fake_stdin):
                result = self.loader.load_file(path, sheet_name="Backup")

            self.assertEqual(result["sheet_name"], "Backup")
            self.assertEqual(result["sheet_names"], ["Visible", "Backup"])
            self.assertEqual(result["dataframe"].shape, (1, 2))
            self.assertEqual(result["dataframe"].iloc[0].to_dict(), {"name": "Grace", "score": "11"})

    def test_multisheet_xlsx_can_consolidate_explicitly(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "multi.xlsx"
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Q1"
            ws1.append(["name", "score"])
            ws1.append(["Ada", 10])
            ws2 = wb.create_sheet("Q2")
            ws2.append(["name", "score"])
            ws2.append(["Grace", 11])
            wb.save(path)

            fake_stdin = mock.Mock()
            fake_stdin.isatty.return_value = False

            with mock.patch.object(self.loader.sys, "stdin", fake_stdin):
                result = self.loader.load_file(path, consolidate_sheets=True)

            self.assertEqual(result["sheet_name"], "[all 2 sheets]")
            self.assertEqual(result["sheet_names"], ["Q1", "Q2"])
            self.assertEqual(result["dataframe"].shape, (2, 2))
            self.assertIn("Consolidated 2 sheets into one table.", result["warnings"])


class LoaderPublicCorpusTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.loader = load_loader_module()
        cls.csv_path = public_fixture(
            "/tmp/realworld-messy-dataset/messy_IMDB_dataset.csv",
        )
        cls.tsv_path = public_fixture(
            "/tmp/wikitablequestions/data/training.tsv",
        )
        cls.xlsx_path = public_fixture(
            "/tmp/calamine-repo/tests/any_sheets.xlsx",
        )
        cls.xls_path = public_fixture(
            "/tmp/calamine-repo/tests/any_sheets.xls",
        )
        cls.xlsm_path = public_fixture(
            "/tmp/calamine-repo/tests/issue221.xlsm",
        )
        cls.ods_path = public_fixture(
            "/tmp/calamine-repo/tests/any_sheets.ods",
        )
        cls.json_path = public_fixture(
            "/tmp/vega-datasets/data/movies.json",
        )
        cls.json_nested_path = public_fixture(
            "/tmp/vega-datasets/data/miserables.json",
        )
        cls.jsonl_path = public_fixture(
            "/tmp/jsonl-repo/tests/data/foo.jsonl",
        )
        cls.corrupt_xls_path = public_fixture(
            "/tmp/xlrd-repo/tests/samples/corrupted_error.xls",
        )

    def test_public_csv_loads_with_semicolon_detection(self):
        if self.csv_path is None:
            self.skipTest("public CSV fixture not available")

        result = self.loader.load_file(self.csv_path)

        self.assertEqual(result["detected_format"], "csv")
        self.assertEqual(result["delimiter"], ";")
        self.assertEqual(result["dataframe"].shape, (101, 12))

    def test_public_tsv_loads(self):
        if self.tsv_path is None:
            self.skipTest("public TSV fixture not available")

        result = self.loader.load_file(self.tsv_path)

        self.assertEqual(result["detected_format"], "tsv")
        self.assertEqual(result["delimiter"], "\t")
        self.assertEqual(result["dataframe"].shape[1], 4)

    def test_public_xlsx_requires_sheet_name_noninteractive(self):
        if self.xlsx_path is None:
            self.skipTest("public XLSX fixture not available")

        with self.assertRaisesRegex(ValueError, "Available sheets"):
            self.loader.load_file(self.xlsx_path)

    def test_public_xlsx_loads_selected_sheet(self):
        if self.xlsx_path is None:
            self.skipTest("public XLSX fixture not available")

        result = self.loader.load_file(self.xlsx_path, sheet_name="Visible")

        self.assertEqual(result["sheet_name"], "Visible")
        self.assertEqual(result["sheet_names"], ["Visible", "Hidden", "VeryHidden"])
        self.assertEqual(result["dataframe"].shape, (4, 2))

    def test_public_xls_loads_selected_sheet(self):
        if self.xls_path is None:
            self.skipTest("public XLS fixture not available")
        if not _XLRD_AVAILABLE:
            self.skipTest("xlrd not installed — run: pip install xlrd")

        result = self.loader.load_file(self.xls_path, sheet_name="Visible")

        self.assertEqual(result["detected_format"], "xls")
        self.assertEqual(result["dataframe"].shape, (4, 2))

    def test_public_xlsm_loads(self):
        if self.xlsm_path is None:
            self.skipTest("public XLSM fixture not available")

        result = self.loader.load_file(self.xlsm_path)

        self.assertEqual(result["detected_format"], "xlsm")
        self.assertEqual(result["sheet_name"], "Sheet1")
        self.assertEqual(result["dataframe"].shape, (1, 2))

    def test_public_ods_loads_selected_sheet(self):
        if self.ods_path is None:
            self.skipTest("public ODS fixture not available")
        if not _ODFPY_AVAILABLE:
            self.skipTest("odfpy not installed — run: pip install odfpy")

        result = self.loader.load_file(self.ods_path, sheet_name="Visible")

        self.assertEqual(result["detected_format"], "ods")
        self.assertEqual(result["sheet_name"], "Visible")
        self.assertEqual(result["dataframe"].shape, (4, 2))

    def test_public_json_loads(self):
        if self.json_path is None:
            self.skipTest("public JSON fixture not available")

        result = self.loader.load_file(self.json_path)

        self.assertEqual(result["detected_format"], "json")
        self.assertEqual(result["dataframe"].shape, (3201, 16))

    def test_public_nested_json_flattens(self):
        if self.json_nested_path is None:
            self.skipTest("public nested JSON fixture not available")

        result = self.loader.load_file(self.json_nested_path)

        self.assertEqual(result["detected_format"], "json")
        self.assertEqual(result["dataframe"].shape, (77, 3))
        self.assertIn("Nested JSON: used array at top-level key 'nodes'", result["warnings"])

    def test_public_jsonl_loads(self):
        if self.jsonl_path is None:
            self.skipTest("public JSONL fixture not available")

        result = self.loader.load_file(self.jsonl_path)

        self.assertEqual(result["detected_format"], "jsonl")
        self.assertEqual(result["dataframe"].shape, (4, 2))

    def test_public_corrupt_xls_raises_clear_error(self):
        if self.corrupt_xls_path is None:
            self.skipTest("public corrupt XLS fixture not available")
        if not _XLRD_AVAILABLE:
            self.skipTest("xlrd not installed — run: pip install xlrd")

        with self.assertRaisesRegex(ValueError, "Could not open workbook"):
            self.loader.load_file(self.corrupt_xls_path)

    def test_public_corrupt_xls_does_not_leak_parser_noise(self):
        if self.corrupt_xls_path is None:
            self.skipTest("public corrupt XLS fixture not available")
        if not _XLRD_AVAILABLE:
            self.skipTest("xlrd not installed — run: pip install xlrd")

        stdout = io.StringIO()
        stderr = io.StringIO()
        with redirect_stdout(stdout), redirect_stderr(stderr):
            with self.assertRaisesRegex(ValueError, "Could not open workbook"):
                self.loader.load_file(self.corrupt_xls_path)

        self.assertEqual(stdout.getvalue(), "")
        self.assertEqual(stderr.getvalue(), "")


if __name__ == "__main__":
    unittest.main()
